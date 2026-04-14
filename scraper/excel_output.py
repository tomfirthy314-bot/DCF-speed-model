"""
Phase 2 — Excel output builder.

Produces a multi-tab workbook:
  Tab 1: Overview         — company info, market data, point-in-time ratios
  Tab 2: Income Statement — historical by year, all scraped sources
  Tab 3: Balance Sheet    — historical by year, all scraped sources
  Tab 4: Cash Flow        — historical by year, all scraped sources

Where two sources report the same metric for the same year, both values are
shown on separate sub-rows labelled by source so discrepancies are visible.

Calculated fields (NOPAT, net debt, margins, WACC etc.) are listed as
placeholder rows — labelled "→ Calculated in model" — ready for Excel formulas
in Phase 3.

Analyst input fields (risk-free rate, ERP, maintenance capex, terminal growth
rate, exit multiple) are listed as blank yellow input rows.
"""

import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
NAVY        = "1F3864"
BLUE        = "2E75B6"
LIGHT_BLUE  = "BDD7EE"
AMBER       = "FFF2CC"
AMBER_DARK  = "FFD966"
LIGHT_GREY  = "F2F2F2"
MID_GREY    = "D9D9D9"
WHITE       = "FFFFFF"
GREEN_LIGHT = "E2EFDA"
RED_LIGHT   = "FCE4D6"


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def build_excel(data: dict, output_dir: str = ".") -> str:
    """
    Build the Excel workbook from the data engine output dict.
    Returns the path to the saved file.
    """
    ticker       = data["ticker"]
    company_name = data["company_name"]
    stats        = data["stats"]
    reconciled   = data["reconciled"]
    raw          = data["raw"]

    date_str  = datetime.today().strftime("%Y-%m-%d")
    safe_name = company_name.replace(" ", "_").replace("/", "-")[:30]
    filename  = os.path.join(output_dir, f"{safe_name}_{ticker}_{date_str}.xlsx")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Sorted years, most recent first
    all_years = sorted(reconciled.keys(), reverse=True)

    _build_overview_tab(wb, ticker, company_name, stats, date_str)
    _build_financials_tab(wb, "Income Statement", IS_METRICS,    reconciled, raw, all_years)
    _build_financials_tab(wb, "Balance Sheet",    BS_METRICS,    reconciled, raw, all_years)
    _build_financials_tab(wb, "Cash Flow",        CF_METRICS,    reconciled, raw, all_years)

    wb.save(filename)
    return filename


# ---------------------------------------------------------------------------
# Metric definitions
# Each entry: (display_label, data_key, row_type)
# row_type: "metric"   → scraped data row
#           "calc"     → calculated in model (placeholder)
#           "input"    → analyst manual input (blank yellow)
#           "section"  → bold section header
#           "spacer"   → empty row
# ---------------------------------------------------------------------------

IS_METRICS = [
    ("REVENUE & PROFIT",         None,                    "section"),
    ("Revenue",                  "revenue",               "metric"),
    ("Gross Profit",             "gross_profit",          "metric"),
    ("Gross Margin %",           None,                    "calc"),
    ("Operating Income",         "operating_income",      "metric"),
    ("Operating Costs (derived)","_op_costs",             "calc"),
    ("EBITDA",                   "ebitda",                "metric"),
    ("EBITDA Margin %",          None,                    "calc"),
    ("Depreciation & Amort.",    "depreciation_amort",    "metric"),
    ("EBIT",                     "ebit",                  "metric"),
    ("EBIT Margin %",            None,                    "calc"),
    ("",                         None,                    "spacer"),
    ("TAX & NET INCOME",         None,                    "section"),
    ("Interest Expense",         "interest_expense",      "metric"),
    ("Pre-Tax Income",           "pre_tax_income",        "metric"),
    ("Tax Provision",            "tax_provision",         "metric"),
    ("Effective Tax Rate %",     None,                    "calc"),
    ("Net Income",               "net_income",            "metric"),
    ("Net Profit Margin %",      "net_profit_margin_pct", "metric"),
    ("EPS (diluted)",            "eps_diluted",           "metric"),
    ("",                         None,                    "spacer"),
    ("NOPAT",                    None,                    "calc"),
]

BS_METRICS = [
    ("ASSETS",                   None,                    "section"),
    ("Total Assets",             "total_assets",          "metric"),
    ("Current Assets",           "current_assets",        "metric"),
    ("Accounts Receivable",      "accounts_receivable",   "metric"),
    ("Inventory",                "inventory",             "metric"),
    ("Cash & Equivalents",       "cash_and_equivalents",  "metric"),
    ("",                         None,                    "spacer"),
    ("LIABILITIES",              None,                    "section"),
    ("Current Liabilities",      "current_liabilities",   "metric"),
    ("Accounts Payable",         "accounts_payable",      "metric"),
    ("Deferred Revenue",         "deferred_revenue",      "metric"),
    ("Total Debt",               "total_debt",            "metric"),
    ("Long-Term Debt",           "long_term_debt",        "metric"),
    ("Lease Liabilities",        "lease_liabilities",     "metric"),
    ("",                         None,                    "spacer"),
    ("EQUITY",                   None,                    "section"),
    ("Total Equity",             "total_equity",          "metric"),
    ("Retained Earnings",        "retained_earnings",     "metric"),
    ("",                         None,                    "spacer"),
    ("DERIVED",                  None,                    "section"),
    ("Net Debt",                 None,                    "calc"),
    ("Capital Structure — Debt %",  None,                 "calc"),
    ("Capital Structure — Equity %",None,                 "calc"),
    ("Change in NWC",            "change_in_working_cap", "metric"),
]

CF_METRICS = [
    ("CASH FLOW",                None,                    "section"),
    ("Operating Cash Flow",      "operating_cash_flow",   "metric"),
    ("Capital Expenditure",      "capex",                 "metric"),
    ("Free Cash Flow",           "free_cash_flow",        "metric"),
    ("Change in Working Capital","change_in_working_cap", "metric"),
    ("",                         None,                    "spacer"),
    ("DERIVED",                  None,                    "section"),
    ("FCFF (Free Cash Flow to Firm)", None,               "calc"),
    ("",                         None,                    "spacer"),
    ("DCF ASSUMPTIONS",          None,                    "section"),
    ("Risk-Free Rate",           None,                    "input"),
    ("Equity Risk Premium",      None,                    "input"),
    ("Cost of Debt",             None,                    "calc"),
    ("WACC",                     None,                    "calc"),
    ("Maintenance CapEx",        None,                    "input"),
    ("Growth CapEx",             None,                    "input"),
    ("Terminal Growth Rate",     None,                    "input"),
    ("Exit Multiple",            None,                    "input"),
]


# ---------------------------------------------------------------------------
# Tab builders
# ---------------------------------------------------------------------------

def _build_overview_tab(wb, ticker, company_name, stats, date_str):
    ws = wb.create_sheet("Overview")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 28

    # Title block
    _write_title(ws, 1, f"{company_name}  ({ticker})", span=2)
    _write_title(ws, 2, f"Data pulled: {date_str}", span=2, small=True)
    ws.append([])

    next_row = _write_legend(ws, start_row=4, col_count=2)
    ws.append([])

    sections = [
        ("COMPANY INFORMATION", [
            ("Ticker",          stats.get("ticker")),
            ("Company Name",    stats.get("company_name")),
            ("Currency",        stats.get("currency")),
            ("Exchange",        stats.get("exchange")),
            ("Sector",          stats.get("sector")),
            ("Industry",        stats.get("industry")),
            ("Country",         stats.get("country")),
        ]),
        ("MARKET DATA", [
            ("Current Price",       _fmt(stats.get("current_price"))),
            ("Market Cap",          _fmt(stats.get("market_cap"))),
            ("Shares Outstanding",  _fmt(stats.get("shares_outstanding"))),
            ("Shares Float",        _fmt(stats.get("shares_float"))),
            ("Beta",                _fmt_ratio(stats.get("beta"))),
        ]),
        ("VALUATION RATIOS", [
            ("Trailing P/E",        _fmt_ratio(stats.get("trailing_pe"))),
            ("Forward P/E",         _fmt_ratio(stats.get("forward_pe"))),
            ("Price-to-Book",       _fmt_ratio(stats.get("price_to_book"))),
        ]),
        ("PROFITABILITY", [
            ("Gross Margin %",      _fmt_pct(stats.get("gross_margins"))),
            ("Operating Margin %",  _fmt_pct(stats.get("operating_margins"))),
            ("Net Profit Margin %", _fmt_pct(stats.get("profit_margins"))),
            ("Return on Equity",    _fmt_pct(stats.get("return_on_equity"))),
            ("Return on Assets",    _fmt_pct(stats.get("return_on_assets"))),
        ]),
        ("FINANCIAL HEALTH", [
            ("Total Debt",          _fmt(stats.get("total_debt"))),
            ("Cash",                _fmt(stats.get("cash"))),
            ("Debt-to-Equity",      _fmt_ratio(stats.get("debt_to_equity"))),
            ("Current Ratio",       _fmt_ratio(stats.get("current_ratio"))),
        ]),
        ("GROWTH (trailing 12m)", [
            ("Revenue Growth %",    _fmt_pct(stats.get("revenue_growth"))),
            ("Earnings Growth %",   _fmt_pct(stats.get("earnings_growth"))),
        ]),
        ("DIVIDENDS", [
            ("Dividend Yield %",    _fmt_pct(stats.get("dividend_yield"))),
            ("Payout Ratio %",      _fmt_pct(stats.get("payout_ratio"))),
        ]),
    ]

    row = next_row + 1
    for section_title, rows in sections:
        # Section header
        ws.cell(row=row, column=1, value=section_title)
        _style_section_header(ws, row, col_count=2)
        row += 1
        for label, value in rows:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            _style_data_row(ws, row, col_count=2, shade=(row % 2 == 0))
            row += 1
        row += 1  # spacer


def _build_financials_tab(wb, tab_name, metric_defs, reconciled, raw, all_years):
    ws = wb.create_sheet(tab_name)

    # Column widths
    ws.column_dimensions["A"].width = 30   # metric label
    ws.column_dimensions["B"].width = 16   # source label
    for i, _ in enumerate(all_years):
        ws.column_dimensions[get_column_letter(3 + i)].width = 14

    col_count = 2 + len(all_years)

    # Legend at top of tab
    legend_end = _write_legend(ws, start_row=1, col_count=col_count)
    ws.append([])

    # Column header row
    header_row = legend_end + 2
    ws.cell(row=header_row, column=1, value=tab_name)
    ws.cell(row=header_row, column=2, value="Source")
    for i, yr in enumerate(all_years):
        ws.cell(row=header_row, column=3 + i, value=yr)
    _style_header_row(ws, header_row, col_count=col_count)

    current_row = header_row + 1
    for label, key, row_type in metric_defs:

        if row_type == "spacer":
            ws.append([])
            current_row += 1
            continue

        if row_type == "section":
            ws.cell(row=current_row, column=1, value=label)
            _style_section_header(ws, current_row, col_count=2 + len(all_years))
            current_row += 1
            continue

        if row_type == "calc":
            ws.cell(row=current_row, column=1, value=f"  {label}")
            ws.cell(row=current_row, column=2, value="→ Calculated in model")
            _style_calc_row(ws, current_row, col_count=2 + len(all_years))
            current_row += 1
            continue

        if row_type == "input":
            ws.cell(row=current_row, column=1, value=f"  {label}")
            ws.cell(row=current_row, column=2, value="← Analyst input")
            _style_input_row(ws, current_row, col_count=2 + len(all_years))
            current_row += 1
            continue

        # --- metric row ---
        # Gather values per year per source
        source_year_vals = _collect_source_values(key, all_years, reconciled, raw)

        # Determine if there are multiple sources with differing values
        multi_source = len(source_year_vals) > 1

        # Primary row: reconciled values
        ws.cell(row=current_row, column=1, value=f"  {label}")
        primary_source = "reconciled" if multi_source else (
            list(source_year_vals.keys())[0] if source_year_vals else "—"
        )
        ws.cell(row=current_row, column=2, value=primary_source)
        shade = (current_row % 2 == 0)
        for i, yr in enumerate(all_years):
            val = reconciled.get(yr, {}).get(key)
            cell = ws.cell(row=current_row, column=3 + i, value=_fmt_num(val))
            _style_value_cell(cell, shade=shade, is_negative=(val is not None and val < 0))
        _style_label_cell(ws.cell(row=current_row, column=1), shade=shade)
        _style_label_cell(ws.cell(row=current_row, column=2), shade=shade, small=True, grey=True)
        current_row += 1

        # Sub-rows: one per source, only when multi-source
        if multi_source:
            for src_name, yr_vals in source_year_vals.items():
                ws.cell(row=current_row, column=1, value=f"      ↳ {src_name}")
                ws.cell(row=current_row, column=2, value=src_name)
                for i, yr in enumerate(all_years):
                    val = yr_vals.get(yr)
                    cell = ws.cell(row=current_row, column=3 + i, value=_fmt_num(val))
                    _style_source_sub_cell(cell, is_negative=(val is not None and val < 0))
                _style_source_sub_label(ws.cell(row=current_row, column=1))
                _style_source_sub_label(ws.cell(row=current_row, column=2))
                current_row += 1

    # Freeze the header row and label/source columns
    ws.freeze_panes = ws.cell(row=header_row + 1, column=3)


# ---------------------------------------------------------------------------
# Source value collector
# ---------------------------------------------------------------------------

def _collect_source_values(key, all_years, reconciled, raw) -> dict[str, dict[str, float]]:
    """
    For a given metric key, return a dict of {source_name: {year: value}}
    covering all sources that have data for this key.
    Only returned when more than one source has data (for sub-row display).
    """
    if key is None:
        return {}

    sources = {}

    # Yahoo annual
    yahoo_yr = raw.get("yahoo", {}).get("financials_by_year", {})
    yahoo_vals = {yr: yahoo_yr.get(yr, {}).get(key) for yr in all_years}
    yahoo_vals = {yr: v for yr, v in yahoo_vals.items() if v is not None}
    if yahoo_vals:
        sources["Yahoo Finance"] = yahoo_vals

    # EDGAR
    edgar = raw.get("edgar", {})
    if edgar.get("available") and edgar.get("latest_year"):
        ey = edgar["latest_year"]
        ev = edgar.get("financials", {}).get(key)
        if ev is not None:
            sources["SEC EDGAR"] = {ey: ev}

    # Macrotrends
    mt_yr = raw.get("macrotrends", {}).get("financials_by_year", {})
    mt_vals = {yr: mt_yr.get(yr, {}).get(key) for yr in all_years}
    mt_vals = {yr: v for yr, v in mt_vals.items() if v is not None}
    if mt_vals:
        sources["Macrotrends"] = mt_vals

    # Only return multiple sources if there are genuinely >1
    return sources if len(sources) > 1 else {}


# ---------------------------------------------------------------------------
# Legend
# ---------------------------------------------------------------------------

LEGEND_ENTRIES = [
    (WHITE,       "000000", False, "Scraped data",
     "Values pulled directly from Yahoo Finance, SEC EDGAR, or Macrotrends."),
    (LIGHT_BLUE,  "17375E", True,  "↳  Source comparison row",
     "Shown when two or more sources report different values for the same metric. "
     "The primary row above shows the reconciled (winning) value."),
    (GREEN_LIGHT, "375623", True,  "→  Calculated in model",
     "Not scraped — to be computed by Excel formula in the DCF build (Phase 3)."),
    (AMBER,       "7F6000", True,  "←  Analyst input required",
     "Not publicly available. Leave blank until you enter your own assumption "
     "(e.g. risk-free rate, terminal growth rate)."),
]

def _write_legend(ws, start_row: int, col_count: int) -> int:
    """
    Write a colour-coded legend block starting at start_row.
    Returns the last row used.
    """
    # Legend title
    title_cell = ws.cell(row=start_row, column=1, value="TABLE KEY")
    title_cell.font = _font(bold=True, size=9, color=NAVY)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    # Extend title background across all columns
    for c in range(1, col_count + 1):
        ws.cell(row=start_row, column=c).fill = _fill(MID_GREY)
    ws.row_dimensions[start_row].height = 14

    row = start_row + 1
    for bg, text_color, italic, label, description in LEGEND_ENTRIES:
        # Swatch cell (column 1)
        swatch = ws.cell(row=row, column=1, value=f"  {label}")
        swatch.fill = _fill(bg)
        swatch.font = _font(size=9, color=text_color, italic=italic)
        swatch.alignment = Alignment(horizontal="left", vertical="center")

        # Description cell (column 2 onwards, merged visually via wide column)
        desc = ws.cell(row=row, column=2, value=description)
        desc.fill = _fill(bg)
        desc.font = _font(size=9, color="606060", italic=True)
        desc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

        # Fill remaining columns with the same background
        for c in range(3, col_count + 1):
            ws.cell(row=row, column=c).fill = _fill(bg)

        ws.row_dimensions[row].height = 13
        row += 1

    return row  # first empty row after legend


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")

def _border_bottom():
    return Border(bottom=Side(style="thin", color="CCCCCC"))

def _style_header_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill(NAVY)
        cell.font = _font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center" if c > 2 else "left", vertical="center")
    ws.row_dimensions[row].height = 18

def _style_section_header(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill(BLUE)
        cell.font = _font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 16

def _style_data_row(ws, row, col_count, shade=False):
    bg = LIGHT_GREY if shade else WHITE
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill(bg)
        cell.font = _font(size=10)
        cell.alignment = Alignment(horizontal="left" if c == 1 else "right", vertical="center")

def _style_label_cell(cell, shade=False, small=False, grey=False):
    bg = LIGHT_GREY if shade else WHITE
    cell.fill = _fill(bg)
    cell.font = _font(size=9 if small else 10, color="808080" if grey else "000000")
    cell.alignment = Alignment(horizontal="left", vertical="center")

def _style_value_cell(cell, shade=False, is_negative=False):
    bg = LIGHT_GREY if shade else WHITE
    cell.fill = _fill(bg)
    cell.font = _font(size=10, color="C00000" if is_negative else "000000")
    cell.alignment = Alignment(horizontal="right", vertical="center")

def _style_calc_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill(GREEN_LIGHT)
        cell.font = _font(size=9, italic=True, color="375623")
        cell.alignment = Alignment(horizontal="left" if c <= 2 else "right", vertical="center")

def _style_input_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill(AMBER)
        cell.font = _font(size=9, italic=True, color="7F6000")
        cell.alignment = Alignment(horizontal="left" if c <= 2 else "right", vertical="center")

def _style_source_sub_cell(cell, is_negative=False):
    cell.fill = _fill(LIGHT_BLUE)
    cell.font = _font(size=9, color="C00000" if is_negative else "17375E", italic=True)
    cell.alignment = Alignment(horizontal="right", vertical="center")

def _style_source_sub_label(cell):
    cell.fill = _fill(LIGHT_BLUE)
    cell.font = _font(size=9, color="17375E", italic=True)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def _write_title(ws, row, text, span=4, small=False):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _font(bold=not small, size=8 if small else 13, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20 if not small else 14


# ---------------------------------------------------------------------------
# Number formatting helpers
# ---------------------------------------------------------------------------

def _fmt_num(val) -> str | None:
    """Format large numbers for display in cells."""
    if val is None:
        return None
    try:
        val = float(val)
    except (TypeError, ValueError):
        return str(val)
    if abs(val) >= 1e12:
        return f"{val / 1e12:.2f}T"
    if abs(val) >= 1e9:
        return f"{val / 1e9:.2f}B"
    if abs(val) >= 1e6:
        return f"{val / 1e6:.0f}M"
    if abs(val) < 1000:
        return f"{val:.2f}"
    return f"{val:,.0f}"

def _fmt(val) -> str:
    if val is None:
        return "—"
    try:
        val = float(val)
    except (TypeError, ValueError):
        return str(val)
    if abs(val) >= 1e12:
        return f"{val / 1e12:.2f}T"
    if abs(val) >= 1e9:
        return f"{val / 1e9:.2f}B"
    if abs(val) >= 1e6:
        return f"{val / 1e6:.2f}M"
    return f"{val:,.0f}"

def _fmt_ratio(val) -> str:
    if val is None:
        return "—"
    try:
        return f"{float(val):.2f}x"
    except (TypeError, ValueError):
        return str(val)

def _fmt_pct(val) -> str:
    if val is None:
        return "—"
    try:
        v = float(val)
        # yfinance returns margins as decimals (0.25 = 25%)
        if abs(v) <= 1.5:
            v = v * 100
        return f"{v:.1f}%"
    except (TypeError, ValueError):
        return str(val)
