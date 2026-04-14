"""
DCF Excel Model Builder — v0.1

Produces a multi-tab Excel workbook from the full pipeline output.

Layout inspired by CFI DCF model conventions:
  - Blue text for inputs/assumptions
  - Black text for calculated cells
  - Compact assumptions panel alongside main DCF table
  - Equity bridge below DCF calculation
  - Market value vs intrinsic value comparison

Tabs:
  1. DCF Model   — main sheet: assumptions, FCF table, terminal value, equity bridge
  2. Historical  — clean historical financials (IS, CF, BS)
  3. Forecast    — 3-scenario detailed projections
  4. Sensitivity — WACC × terminal growth value-per-share grid
  5. Pipeline    — data quality, stage statuses, confidence scores

v0.1 NOTES:
  - To be redesigned when reference model examples are provided
  - All numbers in reporting currency millions unless noted
  - Charts not yet implemented (planned for v0.2)
  - Live Excel formulas not yet implemented — values written directly from Python
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
C_NAVY   = "1F3864"   # section headers
C_BLUE   = "2E75B6"   # sub-headers
C_INPUT  = "2E75B6"   # input text colour (blue, per convention)
C_HIST   = "DEEAF1"   # historical column tint
C_FORE   = "FFF2CC"   # forecast column tint
C_LGRAY  = "F2F2F2"   # alternating row
C_WHITE  = "FFFFFF"
C_GREEN  = "E2EFDA"   # positive / above market
C_RED    = "FCE4D6"   # negative / below market
C_AMBER  = "FFF2CC"   # caution / assumption default
C_WARN   = "ED7D31"   # warning text
C_DARK   = "1A1A1A"   # near-black for main text

_DIV = 1_000_000      # display in millions
FMT_M = '#,##0.0"M"'  # e.g.  1,234.5M
FMT_B = '#,##0.0"B"'  # e.g.      1.2B  (used when value ≥ 1 billion)


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def build_model(valued: dict, output_dir: str = ".") -> str:
    """Build the full DCF Excel workbook and return the file path."""
    meta = _extract_meta(valued)

    wb = Workbook()
    wb.remove(wb.active)

    _build_explainer_sheet(wb, valued, meta)
    _build_dcf_sheet(wb, valued, meta)
    _build_historical_sheet(wb, valued, meta)
    _build_forecast_sheet(wb, valued, meta)
    _build_sensitivity_sheet(wb, valued, meta)
    _build_wacc_sheet(wb, valued, meta)
    _build_comparables_sheet(wb, valued, meta)
    _build_raw_data_sheet(wb, valued, meta)
    _build_data_audit_sheet(wb, valued, meta)
    _build_pipeline_sheet(wb, valued, meta)

    fname = (f"{meta['safe_name']}_{meta['ticker']}_{meta['date']}_DCF.xlsx")
    path  = os.path.join(output_dir, fname)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Tab 1 — DCF Model (main sheet)
# ---------------------------------------------------------------------------

def _build_dcf_sheet(wb: Workbook, valued: dict, meta: dict):
    ws = wb.create_sheet("DCF Model")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C22"

    norm        = valued.get("normalised", {})
    assumptions = valued.get("assumptions", {})
    forecast    = valued.get("forecast", {})
    valuation   = valued.get("valuation", {})
    wc          = valued.get("wacc_components", {})
    bridge_in   = valued.get("equity_bridge_inputs", {})
    stats       = valued.get("stats", {})
    ccy         = stats.get("currency", "")
    n_years     = valued.get("forecast_years", 5)
    start_yr    = valued.get("forecast_start_year", datetime.utcnow().year + 1)
    base_yr     = valued.get("base_year", "")
    base_proj   = forecast.get("base", {})
    year_keys   = sorted(base_proj.keys())

    # Column setup: A=labels, B=entry/base, C..G=forecast years, H=terminal, I=gap, J-K=TV section
    _col_width(ws, 1, 36)   # A - labels
    for c in range(2, 3 + n_years + 1):   # B (base year) through terminal — all uniform
        _col_width(ws, c, 13)
    _col_width(ws, 3 + n_years + 1, 2)   # gap
    _col_width(ws, 3 + n_years + 2, 22)  # TV label
    _col_width(ws, 3 + n_years + 3, 14)  # TV value

    TC = 3 + n_years + 2   # first col of right panel
    TV = TC + 1             # value col of right panel

    # ====================================================================
    # ROWS 1-7: Title + VPS Spotlight banner
    # ====================================================================
    vps          = valuation.get("base", {}).get("value_per_share")
    price        = stats.get("current_price")
    implied_ret  = valuation.get("base", {}).get("implied_return")
    bear_vps     = valuation.get("bear", {}).get("value_per_share")
    bull_vps     = valuation.get("bull", {}).get("value_per_share")

    is_up   = vps is not None and price is not None and vps >= price
    ret_pct = f"{abs(implied_ret):.1%}" if implied_ret is not None else "—"
    dir_word = "UPSIDE" if is_up else "DOWNSIDE"
    arrow    = "▲" if is_up else "▼"
    C_DIR    = "375623" if is_up else "C00000"      # dark green / dark red text
    C_DIR_BG = "E2EFDA" if is_up else "FCE4D6"      # green / red fill

    r = 1
    # Row 1 — company title
    _merge_write(ws, r, 1, 3 + n_years,
                 f"DCF VALUATION MODEL  —  {meta['company']}  ({meta['ticker']})",
                 font=Font(bold=True, size=13, color=C_WHITE),
                 fill=_fill(C_NAVY), height=24,
                 align=Alignment(horizontal="left", vertical="center", indent=2))

    # Row 2 — spacer strip
    r = 2
    ws.row_dimensions[r].height = 4

    # Row 3 — spotlight labels
    r = 3
    _merge_write(ws, r, 1, 2,
                 f"Implied Share Price ({ccy})",
                 font=Font(bold=True, size=9, color="808080"),
                 fill=_fill("F2F2F2"), height=14,
                 align=Alignment(horizontal="center", vertical="center"))
    _merge_write(ws, r, 3, 4,
                 f"Current Market Price ({ccy})",
                 font=Font(bold=True, size=9, color="808080"),
                 fill=_fill("F2F2F2"), height=14,
                 align=Alignment(horizontal="center", vertical="center"))
    _merge_write(ws, r, 5, 6,
                 f"{arrow}  {dir_word}",
                 font=Font(bold=True, size=9, color=C_DIR),
                 fill=_fill(C_DIR_BG), height=14,
                 align=Alignment(horizontal="center", vertical="center"))
    _merge_write(ws, r, 7, 8,
                 "Confidence",
                 font=Font(bold=True, size=9, color="808080"),
                 fill=_fill("F2F2F2"), height=14,
                 align=Alignment(horizontal="center", vertical="center"))

    # Row 4 — spotlight values (large, yellow highlighted)
    r = 4
    explanation  = valued.get("explanation", {})
    conf         = (explanation.get("overall_confidence") or "—").upper()
    conf_colour  = {"HIGH": "375623", "MEDIUM": "7F6000", "LOW": "C00000"}.get(conf, C_DARK)
    conf_bg      = {"HIGH": "E2EFDA", "MEDIUM": "FFF2CC", "LOW": "FCE4D6"}.get(conf, "F2F2F2")
    C_YELLOW_BG  = "FFFF00"   # matching reference model yellow

    ws.row_dimensions[r].height = 30

    def _spotlight(col1, col2, value, fmt, bg, text_color=C_DARK, size=16):
        ws.merge_cells(start_row=r, start_column=col1, end_row=r, end_column=col2)
        c = ws.cell(row=r, column=col1, value=value)
        c.number_format = fmt
        c.font      = Font(bold=True, size=size, color=text_color)
        c.fill      = _fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")

    _spotlight(1, 2, vps,   "0.00", C_YELLOW_BG)
    _spotlight(3, 4, price, "0.00", C_YELLOW_BG)
    _spotlight(5, 6, implied_ret, "0.0%", C_DIR_BG, text_color=C_DIR)
    _spotlight(7, 8, conf, "@",    conf_bg,  text_color=conf_colour)

    # Row 5 — scenario strip  (Bear | Base | Bull)
    r = 5
    ws.row_dimensions[r].height = 13
    for col, sc_name, sc_vps, bg in [
        (1, "BEAR", bear_vps, "FCE4D6"),
        (3, "BASE", vps,      "DEEAF1"),
        (5, "BULL", bull_vps, "E2EFDA"),
    ]:
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
        txt = f"{sc_name}  {ccy} {sc_vps:.2f}" if sc_vps else sc_name
        c = ws.cell(row=r, column=col, value=txt)
        c.font      = Font(bold=True, size=8, color=C_DARK)
        c.fill      = _fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 6 — disclaimer strip
    r = 6
    _merge_write(ws, r, 1, 3 + n_years,
                 "⚠  v0.1  |  FOR REVIEW ONLY  |  All values in reporting currency millions unless noted",
                 font=Font(italic=True, size=8, color=C_WARN),
                 fill=_fill("FFF2CC"), height=13,
                 align=Alignment(horizontal="left", vertical="center", indent=2))

    # Row 7 — spacer
    r = 7
    ws.row_dimensions[r].height = 6

    # ====================================================================
    # LEFT PANEL — Assumptions (rows 8-17, cols A-B)
    # Compact — 6 key inputs only. Full WACC detail is in the WACC tab.
    # ====================================================================
    r = 8
    _section_header(ws, r, 1, 2, f"Assumptions  |  {meta['company']}  ({ccy} millions)")
    r += 1

    wacc_val = assumptions.get("wacc", {}).get("value")
    tg_val   = assumptions.get("terminal_growth", {}).get("value")
    tax_val  = assumptions.get("tax_rate", {}).get("value")

    rev_g_y1 = assumptions.get("revenue_growth", {}).get("year_1")
    ebit_m_y1 = assumptions.get("ebit_margin", {}).get("year_1")
    ebit_m_y5 = assumptions.get("ebit_margin", {}).get(f"year_{n_years}")

    asm_rows = [
        ("Valuation Date",              meta["date"],                          "@"),
        ("Discount Rate (WACC)",        wacc_val,                              "0.0%"),
        ("Terminal Growth Rate",        tg_val,                                "0.0%"),
        ("Effective Tax Rate",          tax_val,                               "0.0%"),
        ("Revenue Growth (Year 1)",     rev_g_y1,                              "0.0%"),
        (f"EBIT Margin (Yr1 → Yr{n_years})",
                                        f"{ebit_m_y1:.1%} → {ebit_m_y5:.1%}" if ebit_m_y1 and ebit_m_y5 else None,
                                        "@"),
        ("Beta (Blume-adjusted)",       wc.get("beta_adjusted") or wc.get("beta"), "0.00"),
        (f"All figures in {ccy} millions unless noted", None, "@"),
    ]

    for label, val, fmt in asm_rows:
        _label_cell(ws, r, 1, label, italic=(label.startswith("All figures")))
        if val is not None:
            _input_cell(ws, r, 2, val, fmt)
        ws.row_dimensions[r].height = 14
        r += 1

    # ====================================================================
    # RIGHT PANEL — TV Summary + EV Bridge (rows 8-17, cols TC-TV)
    # ====================================================================
    r = 8
    base_val = valuation.get("base", {})
    _section_header(ws, r, TC, TV, "Enterprise Value Summary")
    r += 1

    ev    = base_val.get("enterprise_value")
    br    = base_val.get("equity_bridge", {})
    vps   = base_val.get("value_per_share")
    price = stats.get("current_price")

    ev_v,  ev_fmt  = _smart_m(ev)
    pv_tv_v, pv_tv_fmt = _smart_m(base_val.get("pv_terminal_value"))
    pv_fc_v, pv_fc_fmt = _smart_m(base_val.get("pv_forecast_period"))
    csh_v, csh_fmt = _smart_m(br.get("plus_cash"))
    dbt_v, dbt_fmt = _smart_m(br.get("less_debt"))
    lse_v, lse_fmt = _smart_m(br.get("less_lease_liabilities"))
    eqv_v, eqv_fmt = _smart_m(br.get("equity_value"))

    def _rp(label, val, fmt, bold=False, bg=None, top=False):
        _label_cell(ws, r, TC, label, bold=bold)
        if val is not None:
            cell = ws.cell(row=r, column=TV, value=val)
            cell.number_format = fmt
            cell.font = Font(bold=bold, size=10, color=C_NAVY if bold else C_DARK)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if bg:
                cell.fill = _fill(bg)
            if top:
                cell.border = _top_border()
        ws.row_dimensions[r].height = 14

    _rp("PV of Forecast Cash Flows",  pv_fc_v,  pv_fc_fmt)
    r += 1
    _rp("PV of Terminal Value",       pv_tv_v,  pv_tv_fmt)
    r += 1
    _rp("Terminal Value % of EV",     base_val.get("terminal_value_pct_ev"), "0.0%")
    r += 1
    _rp("Enterprise Value",           ev_v,     ev_fmt,    bold=True, top=True)
    r += 1
    _rp("Less: Debt",                 dbt_v,    dbt_fmt)
    r += 1
    _rp("Less: Lease Liabilities",    lse_v,    lse_fmt)
    r += 1
    _rp("Plus: Cash",                 csh_v,    csh_fmt)
    r += 1
    _rp("Equity Value",               eqv_v,    eqv_fmt,   bold=True, top=True)
    r += 1
    _rp(f"Implied Share Price ({ccy})", vps,    "0.00",    bold=True, bg="FFFF00")
    r += 1

    # ====================================================================
    # DCF CALCULATION TABLE
    # ====================================================================

    # --- pull base year actual data for historical column ---
    years_data = valued.get("canonical_by_year", {})
    by = years_data.get(base_yr, {})
    prev_yrs = sorted(years_data.keys(), reverse=True)
    prev_yr  = next((y for y in prev_yrs if y < base_yr), None)
    by_prev  = years_data.get(prev_yr, {}) if prev_yr else {}

    def _safe(v):
        return float(v) if v is not None else None

    # Base year actuals
    by_rev   = _safe(by.get("revenue"))
    by_ebitda= _safe(by.get("ebitda"))
    by_da    = _safe(by.get("da"))
    by_ebit  = _safe(by.get("ebit"))
    by_tax   = _safe(by.get("tax_provision"))
    by_nopat = (by_ebit * (1 - tax_val)) if by_ebit is not None and tax_val else None
    by_capex = _safe(by.get("capex"))
    by_ar    = _safe(by.get("accounts_receivable"))
    by_ap    = _safe(by.get("accounts_payable"))
    by_inv   = _safe(by.get("inventory")) or 0
    by_nwc   = (by_ar + by_inv - by_ap) if (by_ar is not None and by_ap is not None) else None
    by_fcff  = _safe(by.get("free_cash_flow"))

    # Prior year revenue for base-year growth
    prev_rev = _safe(by_prev.get("revenue")) if by_prev else None
    by_rev_g = ((by_rev - prev_rev) / abs(prev_rev)) if (by_rev and prev_rev and prev_rev != 0) else None

    # Derived % metrics for base year
    def _pct(num, denom):
        if num is None or not denom:
            return None
        return num / denom

    r_tbl = r + 2
    ws.row_dimensions[r + 1].height = 6
    r = r_tbl

    # Section header
    _section_header(ws, r, 1, 3 + n_years,
                    f"Discounted Cash Flow  |  Base Year: {base_yr}  |  Values in {ccy} (B/M)",
                    color=C_NAVY)
    ws.row_dimensions[r].height = 18

    # Column headers
    r += 1
    ws.row_dimensions[r].height = 16
    _header_cell(ws, r, 1, "")
    _header_cell(ws, r, 2, f"{base_yr}A", bg=C_HIST)
    for i, yk in enumerate(year_keys):
        cal = base_proj[yk]["calendar_year"]
        _header_cell(ws, r, 3 + i, f"{cal}E", bg=C_FORE)
    _header_cell(ws, r, 3 + n_years, "Terminal", bg=C_LGRAY)

    # ---- helpers scoped to table ----

    def _tot(label, bv, fy_vals, fmt, term_val=None, bold=True, bg_hist=C_HIST, bg_fore=C_FORE):
        """Total row — bold, top border."""
        nonlocal r
        r += 1
        ws.row_dimensions[r].height = 16
        _label_cell(ws, r, 1, label, bold=bold)
        bv_d, bv_f = _smart_m(bv) if fmt == "smart" else (_m(bv), fmt)
        _calc_cell(ws, r, 2, bv_d, bv_f, bold=bold, bg=bg_hist, border=_top_border())
        for i, yk in enumerate(year_keys):
            v = fy_vals[i]
            vd, vf = _smart_m(v) if fmt == "smart" else (_m(v), fmt)
            _calc_cell(ws, r, 3 + i, vd, vf, bold=bold, bg=bg_fore, border=_top_border())
        if term_val is not None:
            tv_d, tv_f = _smart_m(term_val) if fmt == "smart" else (_m(term_val), fmt)
            _calc_cell(ws, r, 3 + n_years, tv_d, tv_f, bold=bold,
                       bg=C_LGRAY, border=_top_border())

    def _dat(label, bv, fy_vals, fmt=FMT_M, bold=False):
        """Plain data row."""
        nonlocal r
        r += 1
        ws.row_dimensions[r].height = 15
        _label_cell(ws, r, 1, label, bold=bold)
        bv_d, bv_f = _smart_m(bv) if fmt == "smart" else (_m(bv), fmt)
        _calc_cell(ws, r, 2, bv_d, bv_f, bold=bold, bg=C_HIST)
        for i, yk in enumerate(year_keys):
            v = fy_vals[i]
            vd, vf = _smart_m(v) if fmt == "smart" else (_m(v), fmt)
            _calc_cell(ws, r, 3 + i, vd, vf, bold=bold, bg=C_FORE)

    def _sub(label, bv, fy_vals, fmt="0.0%"):
        """Sub-row — italic, indented, lighter text."""
        nonlocal r
        r += 1
        ws.row_dimensions[r].height = 13
        cell = ws.cell(row=r, column=1, value=f"  {label}")
        cell.font      = Font(italic=True, size=9, color="606060")
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        _calc_cell(ws, r, 2, bv, fmt, italic=True, bg=C_HIST, color="606060")
        for i, yk in enumerate(year_keys):
            v = fy_vals[i]
            _calc_cell(ws, r, 3 + i, v, fmt, italic=True, bg=C_FORE, color="606060")

    def _gap(h=4):
        nonlocal r
        r += 1
        ws.row_dimensions[r].height = h

    def _fy(field):
        return [base_proj[yk].get(field) for yk in year_keys]

    def _fy_pct(num_field, denom_field):
        result = []
        for yk in year_keys:
            n = base_proj[yk].get(num_field)
            d = base_proj[yk].get(denom_field)
            result.append(_pct(n, d))
        return result

    # ── Revenue ──────────────────────────────────────────────────────────────
    _gap()
    fy_rev   = _fy("revenue")
    fy_rev_g = [base_proj[yk].get("revenue_growth") for yk in year_keys]
    _tot("Revenue", by_rev, fy_rev, "smart", bold=True)
    _sub("% Growth", by_rev_g, fy_rev_g)

    # ── EBITDA ───────────────────────────────────────────────────────────────
    _gap()
    fy_ebitda = [
        (base_proj[yk].get("ebit") or 0) + (base_proj[yk].get("da") or 0)
        for yk in year_keys
    ]
    fy_ebitda_m = [_pct(v, base_proj[yk].get("revenue"))
                   for v, yk in zip(fy_ebitda, year_keys)]
    _dat("EBITDA", by_ebitda, fy_ebitda, fmt="smart")
    _sub("% Margin", _pct(by_ebitda, by_rev), fy_ebitda_m)

    # ── D&A ──────────────────────────────────────────────────────────────────
    fy_da = _fy("da")
    _dat("Depreciation & Amortisation", by_da, fy_da, fmt="smart")
    _sub("% Revenue", _pct(by_da, by_rev),
         [_pct(d, base_proj[yk].get("revenue")) for d, yk in zip(fy_da, year_keys)])

    # ── EBIT ─────────────────────────────────────────────────────────────────
    _gap()
    fy_ebit   = _fy("ebit")
    fy_ebit_m = _fy_pct("ebit", "revenue")
    _tot("EBIT", by_ebit, fy_ebit, "smart")
    _sub("% Margin", _pct(by_ebit, by_rev), fy_ebit_m)

    # ── Tax ──────────────────────────────────────────────────────────────────
    _gap()
    fy_tax = _fy("tax")
    fy_tax_r = [tax_val] * len(year_keys)
    _dat("Tax", by_tax, fy_tax, fmt="smart")
    _sub("Tax Rate", _pct(by_tax, by_ebit) if (by_tax and by_ebit) else None, fy_tax_r)

    # ── NOPAT ────────────────────────────────────────────────────────────────
    fy_nopat   = _fy("nopat")
    fy_nopat_m = _fy_pct("nopat", "revenue")
    _tot("NOPAT", by_nopat, fy_nopat, "smart")
    _sub("% Margin", _pct(by_nopat, by_rev), fy_nopat_m)

    # ── Capex ────────────────────────────────────────────────────────────────
    _gap()
    fy_capex   = _fy("capex")
    fy_capex_p = [_pct(abs(v) if v else None, base_proj[yk].get("revenue"))
                  for v, yk in zip(fy_capex, year_keys)]
    _dat("Capital Expenditure", by_capex, fy_capex, fmt="smart")
    _sub("% Revenue", _pct(abs(by_capex) if by_capex else None, by_rev), fy_capex_p)

    # D&A add-back (same values, shown again per reference convention)
    _dat("Add: Depreciation & Amortisation", by_da, fy_da, fmt="smart")

    # ── NWC ──────────────────────────────────────────────────────────────────
    _gap()
    fy_nwc = [base_proj[yk].get("revenue") * assumptions.get("nwc_pct_revenue", {}).get("value", 0)
              for yk in year_keys]
    fy_nwc_p = [assumptions.get("nwc_pct_revenue", {}).get("value")] * len(year_keys)
    _dat("Net Working Capital", by_nwc, fy_nwc, fmt="smart")
    _sub("% Revenue", _pct(by_nwc, by_rev), fy_nwc_p)

    fy_dnwc = _fy("delta_nwc")
    _dat("Change in Net Working Capital", None, fy_dnwc, fmt="smart")

    # ── Free Cash Flow ────────────────────────────────────────────────────────
    _gap(6)
    fy_fcff   = _fy("fcff")
    fy_fcff_p = _fy_pct("fcff", "revenue")
    tv_raw    = base_val.get("terminal_value")
    _tot("Free Cash Flow", by_fcff, fy_fcff, "smart",
         term_val=tv_raw, bold=True)
    _sub("% Revenue", _pct(by_fcff, by_rev), fy_fcff_p)

    # ── Discount mechanics ───────────────────────────────────────────────────
    _gap()
    r += 1
    ws.row_dimensions[r].height = 14
    cell = ws.cell(row=r, column=1, value="Discount Period")
    cell.font      = Font(italic=True, size=9, color="606060")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    _calc_cell(ws, r, 2, 0, "0", italic=True, color="606060", bg=C_HIST)
    for i in range(n_years):
        _calc_cell(ws, r, 3 + i, i + 1, "0", italic=True, color="606060", bg=C_FORE)

    r += 1
    ws.row_dimensions[r].height = 14
    cell = ws.cell(row=r, column=1, value="Discount Factor")
    cell.font      = Font(italic=True, size=9, color="606060")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    _calc_cell(ws, r, 2, 1.0, "0.000", italic=True, color="606060", bg=C_HIST)
    for i in range(n_years):
        df = 1 / (1 + wacc_val) ** (i + 1) if wacc_val else None
        _calc_cell(ws, r, 3 + i, df, "0.000", italic=True, color="606060", bg=C_FORE)

    # ── Discounted FCF ────────────────────────────────────────────────────────
    _gap(6)
    pv_by_yr = base_val.get("pv_by_year", {})
    fy_pv    = [pv_by_yr.get(yk) for yk in year_keys]
    _tot("Discounted FCF", None, fy_pv, "smart",
         term_val=base_val.get("pv_terminal_value"), bold=True)

    r += 1
    ws.row_dimensions[r].height = 10  # spacer before bridges

    # ====================================================================
    # BOTTOM PANELS — Equity Bridge | Market Value | vs Market | Scenarios
    # ====================================================================
    r += 1
    bridge_start = r

    # ── Intrinsic Value bridge (cols A-B) ────────────────────────────────────
    _section_header(ws, r, 1, 2, f"Intrinsic Value  ({ccy})")
    r += 1

    bridge_rows = [
        ("Enterprise Value",              ev_v,  ev_fmt,  False),
        ("Plus: Cash",                    csh_v, csh_fmt, False),
        ("Less: Debt",                    dbt_v, dbt_fmt, False),
        ("Less: Lease Liabilities",       lse_v, lse_fmt, False),
        ("Equity Value",                  eqv_v, eqv_fmt, True),
        ("",                              None,  "",       False),
        (f"Fully Diluted Shares (M)",
         _m(bridge_in.get("shares_outstanding")), FMT_M,   False),
        (f"Implied Share Price ({ccy})",   vps,   "0.00",  True),
    ]
    for label, val, fmt, bold in bridge_rows:
        _label_cell(ws, r, 1, label, bold=bold)
        cell = ws.cell(row=r, column=2, value=val)
        cell.number_format = fmt
        cell.font = Font(bold=bold, size=10, color=C_NAVY if bold else C_DARK)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if bold and val:
            cell.border   = _top_border()
        if bold and label.startswith("Implied"):
            cell.fill = _fill("FFFF00")
        ws.row_dimensions[r].height = 15
        r += 1

    # ── Market Value bridge (cols D-E) ───────────────────────────────────────
    r = bridge_start
    mc = 4
    _col_width(ws, mc,     22)
    _col_width(ws, mc + 1, 14)
    _section_header(ws, r, mc, mc + 1, f"Market Value  ({ccy})", color=C_BLUE)
    r += 1

    mkt_raw_cap = (price * bridge_in.get("shares_outstanding")) \
                  if price and bridge_in.get("shares_outstanding") else None
    mkt_ev = None
    if mkt_raw_cap and bridge_in.get("debt") is not None and bridge_in.get("cash") is not None:
        mkt_ev = mkt_raw_cap + (bridge_in["debt"] or 0) - (bridge_in["cash"] or 0)

    mc_v,  mc_fmt  = _smart_m(mkt_raw_cap)
    mdt_v, mdt_fmt = _smart_m(bridge_in.get("debt"))
    mcs_v, mcs_fmt = _smart_m(bridge_in.get("cash"))
    mev_v, mev_fmt = _smart_m(mkt_ev)

    mkt_rows = [
        ("Market Capitalisation",    mc_v,  mc_fmt,  False),
        ("Plus: Debt",               mdt_v, mdt_fmt, False),
        ("Less: Cash",               mcs_v, mcs_fmt, False),
        ("Market Enterprise Value",  mev_v, mev_fmt, True),
        ("",                         None,  "",       False),
        ("",                         None,  "",       False),
        (f"Current Share Price ({ccy})", price, "0.00", True),
    ]
    for label, val, fmt, bold in mkt_rows:
        _label_cell(ws, r, mc, label, bold=bold)
        cell = ws.cell(row=r, column=mc + 1, value=val)
        cell.number_format = fmt
        cell.font = Font(bold=bold, size=10, color=C_NAVY if bold else C_DARK)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if bold and val:
            cell.border = _top_border()
        if bold and label.startswith("Current Share"):
            cell.fill = _fill("FFFF00")
        ws.row_dimensions[r].height = 15
        r += 1

    # ── Implied vs Market (cols G-H) ─────────────────────────────────────────
    r = bridge_start
    vc = 7
    _col_width(ws, vc,     22)
    _col_width(ws, vc + 1, 14)
    _section_header(ws, r, vc, vc + 1, "Implied vs Market", color=C_BLUE)
    r += 1

    implied_return = base_val.get("implied_return")
    updown = (vps - price) if (vps and price) else None
    is_pos = updown is not None and updown > 0

    vm_rows = [
        (f"Implied Share Price ({ccy})",     vps,           "0.00", True,  "FFFF00"),
        (f"Current Share Price ({ccy})",     price,         "0.00", True,  "FFFF00"),
        (f"Upside / (Downside) ({ccy})",     updown,        "0.00", True,
         "E2EFDA" if is_pos else "FCE4D6"),
        ("Implied Return",                   implied_return,"0.0%", True,
         "E2EFDA" if is_pos else "FCE4D6"),
        ("",  None, "", False, None),
        ("",  None, "", False, None),
        ("",  None, "", False, None),
    ]
    for label, val, fmt, bold, bg in vm_rows:
        _label_cell(ws, r, vc, label, bold=bold)
        cell = ws.cell(row=r, column=vc + 1, value=val)
        cell.number_format = fmt
        is_pos_v = val is not None and isinstance(val, (int, float)) and val > 0
        is_neg_v = val is not None and isinstance(val, (int, float)) and val < 0
        if bg:
            cell.fill = _fill(bg)
            cell.font = Font(bold=True, size=10 if fmt != "0.00" else 11,
                             color="375623" if is_pos_v else "C00000" if is_neg_v else C_DARK)
        else:
            cell.font = Font(size=10, color=C_DARK)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[r].height = 15
        r += 1

    # ── Scenario Comparison ───────────────────────────────────────────────────
    r = bridge_start + 10
    _section_header(ws, r, 1, 8, "Scenario Comparison", color=C_NAVY)
    r += 1

    _header_cell(ws, r, 1, "")
    for i, sc_name in enumerate(["BEAR", "BASE", "BULL"]):
        bg = C_RED if sc_name == "BEAR" else C_HIST if sc_name == "BASE" else C_GREEN
        _header_cell(ws, r, 2 + i * 2,     sc_name,  bg=bg)
        _header_cell(ws, r, 3 + i * 2, f"vs Market", bg=bg)
    ws.row_dimensions[r].height = 16

    scenario_rows = [
        ("Enterprise Value",             "enterprise_value",               "smart"),
        ("Equity Value",                 ("equity_bridge", "equity_value"), "smart"),
        (f"Implied Share Price ({ccy})", "value_per_share",                "0.00"),
    ]
    for sc_label, field, fmt in scenario_rows:
        r += 1
        ws.row_dimensions[r].height = 15
        bold = (fmt == "0.00")
        _label_cell(ws, r, 1, sc_label, bold=bold)
        for i, sc_name in enumerate(["bear", "base", "bull"]):
            sv = valuation.get(sc_name, {})
            val = sv.get(field[0], {}).get(field[1]) \
                  if isinstance(field, tuple) else sv.get(field)
            val_d, val_f = _smart_m(val) if fmt == "smart" else (val, fmt)
            bg = C_RED if sc_name == "bear" else C_HIST if sc_name == "base" else C_GREEN

            cell = ws.cell(row=r, column=2 + i * 2, value=val_d)
            cell.number_format = val_f
            cell.font = Font(bold=bold, size=10, color=C_DARK)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.fill = _fill(bg)

            if isinstance(field, str) and field == "value_per_share" and price and val_d:
                diff = val_d - price
                dc   = ws.cell(row=r, column=3 + i * 2, value=diff)
                dc.number_format = "0.00"
                dc.font = Font(bold=True, size=10,
                               color="375623" if diff > 0 else "C00000")
                dc.alignment = Alignment(horizontal="right", vertical="center")
                dc.fill = _fill(bg)


# ---------------------------------------------------------------------------
# Tab 1 — Overview (company brief + DCF conclusions)
# ---------------------------------------------------------------------------

def _build_explainer_sheet(wb: Workbook, valued: dict, meta: dict):
    ws = wb.create_sheet("Overview")
    ws.sheet_view.showGridLines = False

    # Column layout: A label-left | B value-left | C gap | D label-right | E value-right | F-G scenario cols
    _col_width(ws, 1, 26)   # A labels
    _col_width(ws, 2, 18)   # B values
    _col_width(ws, 3,  2)   # C gap
    _col_width(ws, 4, 26)   # D labels
    _col_width(ws, 5, 16)   # E values
    _col_width(ws, 6, 12)   # F BEAR
    _col_width(ws, 7, 12)   # G BASE
    _col_width(ws, 8, 12)   # H BULL

    NC = 8   # total columns used

    expl  = valued.get("explanation", {})
    coh   = valued.get("coherence", {})
    stats = valued.get("stats", {})
    norm  = valued.get("normalised", {})
    wc    = valued.get("wacc_components", {})
    val   = valued.get("valuation", {})
    asmps = valued.get("assumptions", {})

    ccy        = stats.get("currency", "")
    base_val   = val.get("base", {})
    vps        = base_val.get("value_per_share")
    ev         = base_val.get("enterprise_value")
    tv_pct     = base_val.get("terminal_value_pct_ev")
    impl_ret   = base_val.get("implied_return")
    price      = stats.get("current_price")
    mktcap     = stats.get("market_cap")
    conf       = expl.get("overall_confidence", "low").upper()

    conf_bg = {"HIGH": "E2EFDA", "MEDIUM": "FFF2CC", "LOW": "FCE4D6"}.get(conf, C_LGRAY)
    conf_fg = {"HIGH": "375623", "MEDIUM": "7F6000", "LOW": "C00000"}.get(conf, C_DARK)
    n_flags = len(coh.get("flags", []))
    n_warns = len(coh.get("warns", []))

    # ---- helpers scoped to this sheet ----

    def _full(r, text, bg=C_NAVY, fg=C_WHITE, bold=True, size=10, height=17, italic=False):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        c = ws.cell(row=r, column=1, value=text)
        c.font      = Font(bold=bold, italic=italic, size=size, color=fg)
        c.fill      = _fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2, wrap_text=True)
        ws.row_dimensions[r].height = height

    def _half_hdr(r, col_start, col_end, text, color=C_BLUE):
        ws.merge_cells(start_row=r, start_column=col_start, end_row=r, end_column=col_end)
        c = ws.cell(row=r, column=col_start, value=text)
        c.font      = Font(bold=True, size=10, color=C_WHITE)
        c.fill      = _fill(color)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 16

    def _kv(r, col_l, col_v, label, value, bold_v=False, fmt=None, color_v=C_DARK, bg=None):
        lc = ws.cell(row=r, column=col_l, value=label)
        lc.font      = Font(size=9, color="595959")
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        vc = ws.cell(row=r, column=col_v, value=value)
        vc.font      = Font(bold=bold_v, size=10, color=color_v)
        vc.alignment = Alignment(horizontal="right", vertical="center")
        if fmt:
            vc.number_format = fmt
        if bg:
            lc.fill = _fill(bg)
            vc.fill = _fill(bg)
        ws.row_dimensions[r].height = 15

    def _spacer(r, h=5):
        ws.row_dimensions[r].height = h

    def _wrap_row(r, col_start, col_end, text, bold=False, italic=False,
                  bg=C_WHITE, fg=C_DARK, size=9, min_h=15):
        ws.merge_cells(start_row=r, start_column=col_start, end_row=r, end_column=col_end)
        c = ws.cell(row=r, column=col_start, value=text)
        c.font      = Font(bold=bold, italic=italic, size=size, color=fg)
        c.fill      = _fill(bg)
        c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ws.row_dimensions[r].height = max(min_h, min(80, size * (len(text) // 85 + 1) + 4))

    def _sm(v):
        """Smart M/B format value for display string."""
        if v is None: return "—"
        v = float(v)
        if abs(v) >= 1e9:  return f"{v/1e9:.1f}B"
        if abs(v) >= 1e6:  return f"{v/1e6:.0f}M"
        return f"{v:,.0f}"

    # ======================================================================
    # ROW 1-2: Title + confidence banner
    # ======================================================================
    r = 1
    _full(r, f"DCF VALUATION  —  {meta['company']}  ({meta['ticker']})",
          height=26, size=13)
    r += 1
    dq_score = valued.get("data_quality", {}).get("quality_score", "—")
    status_text = (
        f"Run: {meta['date']}   |   Currency: {ccy}   |   "
        f"Confidence: {conf}   |   {n_flags} flag(s)  ·  {n_warns} warning(s)   |   "
        f"Data quality: {dq_score}/100"
    )
    _full(r, status_text, bg=conf_bg, fg=conf_fg, bold=False, size=9, height=15)
    r += 1; _spacer(r); r += 1

    # ======================================================================
    # ROW 4-10: Two-panel — Company Profile (left) | DCF Conclusions (right)
    # ======================================================================
    _half_hdr(r, 1, 2, "COMPANY PROFILE")
    _half_hdr(r, 4, 8, "DCF CONCLUSIONS")
    r += 1

    # Left: company profile rows
    profile = [
        ("Sector",          stats.get("sector", "—")),
        ("Industry",        stats.get("industry", "—")),
        ("Country",         stats.get("country", "—")),
        ("Classification",  valued.get("classification", "—").replace("_", " ").title()),
        ("Base Year",       valued.get("base_year", "—")),
        ("Exchange",        stats.get("exchange", "—")),
    ]

    # Right: DCF conclusions
    direction = "below" if (vps and price and vps < price) else "above"
    implied_pct = f"{abs(impl_ret):.1%} {direction} market" if impl_ret else "—"
    wacc_val = (valued.get("assumptions", {}).get("wacc", {}) or {}).get("value")
    tgr_val  = (valued.get("assumptions", {}).get("terminal_growth", {}) or {}).get("value")
    conclusions = [
        ("Base VPS",         f"{ccy} {vps:.2f}"    if vps   else "—", "C00000" if (vps and price and vps < price) else "375623"),
        ("Market Price",     f"{ccy} {price:.2f}"  if price else "—", C_DARK),
        ("Implied Return",   implied_pct,                               conf_fg),
        ("Enterprise Value", _sm(ev),                                   C_DARK),
        ("WACC",             f"{wacc_val:.1%}"      if wacc_val else "—", C_DARK),
        ("Terminal Growth",  f"{tgr_val:.1%}"       if tgr_val  else "—", C_DARK),
    ]

    max_rows = max(len(profile), len(conclusions))
    for i in range(max_rows):
        if i < len(profile):
            _kv(r, 1, 2, profile[i][0], profile[i][1])
        if i < len(conclusions):
            lbl, val_str, col = conclusions[i]
            _kv(r, 4, 5, lbl, val_str, bold_v=(i in (0, 2)), color_v=col)
        r += 1

    _spacer(r); r += 1

    # ======================================================================
    # ROW: Financial Snapshot (left) | Scenario Comparison (right)
    # ======================================================================
    _half_hdr(r, 1, 2, "FINANCIAL SNAPSHOT  (" + ccy + " M)")
    _half_hdr(r, 4, 8, "SCENARIO COMPARISON  (" + ccy + ")")
    r += 1

    # Financial snapshot from normalised base year
    revenue  = norm.get("revenue")
    ebit     = norm.get("ebit")
    ebit_m   = norm.get("ebit_margin")
    da       = norm.get("da")
    capex    = norm.get("capex")
    base_yr  = valued.get("base_year", "")

    fin_rows = [
        ("Revenue",         _sm(revenue)),
        ("EBIT",            _sm(ebit)),
        ("EBIT Margin",     f"{ebit_m:.1%}"   if ebit_m  else "—"),
        ("D&A",             _sm(da)),
        ("Capex",           _sm(capex)),
        ("Market Cap",      _sm(mktcap)),
        ("Current Price",   f"{ccy} {price:.2f}" if price else "—"),
        ("Beta",            f"{stats.get('beta', '—'):.2f}" if stats.get("beta") else "—"),
    ]

    # Scenario headers
    sc_label_col = [(6, "BEAR", C_RED), (7, "BASE", C_HIST), (8, "BULL", C_GREEN)]
    for col, lbl, bg in sc_label_col:
        c = ws.cell(row=r, column=col, value=lbl)
        c.font      = Font(bold=True, size=9, color=C_DARK)
        c.fill      = _fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=r, column=4, value="Metric").font = Font(size=9, color="595959")
    ws.cell(row=r, column=5, value="").font        = Font(size=9)
    ws.row_dimensions[r].height = 14
    r += 1

    sc_rows = [
        ("Enterprise Value",  "enterprise_value",              True),
        ("Equity Value",      ("equity_bridge", "equity_value"), True),
        (f"Per Share ({ccy})", "value_per_share",               False),
    ]

    fin_r = r
    for label, field, use_smart in sc_rows:
        _kv(r, 4, 5, label, "")
        for col, sc_name in [(6, "bear"), (7, "base"), (8, "bull")]:
            sv = val.get(sc_name, {})
            raw = sv.get(field[0], {}).get(field[1]) if isinstance(field, tuple) else sv.get(field)
            if use_smart:
                disp_v, fmt = _smart_m(raw)
            else:
                disp_v = raw
                fmt    = "0.00"
            c = ws.cell(row=r, column=col, value=disp_v)
            c.number_format = fmt
            c.font          = Font(size=9, color=C_DARK)
            c.alignment     = Alignment(horizontal="right", vertical="center")
            bg = C_RED if col == 6 else C_HIST if col == 7 else C_GREEN
            c.fill          = _fill(bg)
        ws.row_dimensions[r].height = 15
        r += 1

    # Fill remaining financial snapshot rows alongside blank scenario rows
    for i, (lbl, val_str) in enumerate(fin_rows):
        row_r = fin_r + i
        if row_r >= r:
            ws.row_dimensions[row_r].height = 15
        _kv(row_r, 1, 2, lbl, val_str)
    r = max(r, fin_r + len(fin_rows))

    _spacer(r); r += 1

    # ======================================================================
    # Key assumptions — compact table (full width)
    # ======================================================================
    _full(r, "KEY ASSUMPTIONS", height=16)
    r += 1

    # Header row
    for col, txt in [(1, "Assumption"), (2, "Value"), (4, "Method / Source"), (5, "Narrative")]:
        c = ws.cell(row=r, column=col, value=txt)
        c.font      = Font(bold=True, size=9, color=C_WHITE)
        c.fill      = _fill(C_BLUE)
        c.alignment = Alignment(horizontal="left" if col != 2 else "right",
                                vertical="center", indent=1)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=NC)
    ws.row_dimensions[r].height = 14
    r += 1

    for a in expl.get("assumption_narrative", []):
        src_col = ("375623" if "live market" in a["source"] else
                   "595959" if "historical"  in a["source"] else "C00000")
        lc = ws.cell(row=r, column=1, value=a["assumption"])
        lc.font = Font(bold=True, size=9, color=C_DARK)
        lc.alignment = Alignment(vertical="top", indent=1)

        vc = ws.cell(row=r, column=2, value=a["value"])
        vc.font = Font(size=9, color=C_DARK)
        vc.alignment = Alignment(horizontal="right", vertical="top")

        sc = ws.cell(row=r, column=4, value=a["source"])
        sc.font = Font(italic=True, size=8, color=src_col)
        sc.alignment = Alignment(vertical="top", wrap_text=True, indent=1)

        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=NC)
        nc = ws.cell(row=r, column=5, value=a["narrative"])
        nc.font = Font(size=9, color=C_DARK)
        nc.alignment = Alignment(wrap_text=True, vertical="top", indent=1)

        h = max(15, min(60, 9 * (len(a["narrative"]) // 90 + 1) + 4))
        ws.row_dimensions[r].height = h
        alt = C_LGRAY if r % 2 == 0 else C_WHITE
        for col in [1, 2, 4]:
            ws.cell(row=r, column=col).fill = _fill(alt)
        ws.cell(row=r, column=5).fill = _fill(alt)
        r += 1

    _spacer(r); r += 1

    # ======================================================================
    # Comparable Companies
    # ======================================================================
    comps = valued.get("comparables", {})
    _full(r, "COMPARABLE COMPANIES", height=16)
    r += 1

    if not comps or not comps.get("available"):
        reason = (comps or {}).get("reason", "Comparables not available.")
        _wrap_row(r, 1, NC, reason, size=9, italic=True, bg=C_LGRAY, fg="595959")
        r += 1
    else:
        peers  = comps.get("peers", [])
        med    = comps.get("median", {})
        impl   = comps.get("implied_from_peers", {})
        rng    = impl.get("vps_range", [])

        def _mx2(v): return f"{v:.1f}x" if v is not None else "n/a"

        # Peer tickers row
        _wrap_row(r, 1, NC,
                  f"Peers:  {', '.join(p.get('ticker','') for p in peers)}",
                  size=9, bold=True, bg=C_LGRAY, fg=C_DARK, min_h=14)
        r += 1

        # Column headers
        for col, txt in [(1,"Metric"),(2,"Peer Median"),(4,"Target (this co.)"),(6,"vs Target")]:
            c = ws.cell(row=r, column=col, value=txt)
            c.font      = Font(bold=True, size=9, color=C_WHITE)
            c.fill      = _fill(C_BLUE)
            c.alignment = Alignment(horizontal="left" if col==1 else "center",
                                    vertical="center", indent=1 if col==1 else 0)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=NC)
        ws.row_dimensions[r].height = 14
        r += 1

        tgt = comps.get("target_metrics", {})
        mult_rows = [
            ("EV / Revenue",  med.get("ev_revenue"),  tgt.get("ev_revenue")),
            ("EV / EBITDA",   med.get("ev_ebitda"),   tgt.get("ev_ebitda")),
            ("EV / EBIT",     med.get("ev_ebit"),     tgt.get("ev_ebit")),
            ("P / E",         med.get("pe"),          tgt.get("pe")),
        ]
        for i, (lbl, peer_v, tgt_v) in enumerate(mult_rows):
            bg_row = C_LGRAY if i % 2 == 0 else C_WHITE
            ws.cell(row=r, column=1, value=lbl).font = Font(size=9, color=C_DARK)
            ws.cell(row=r, column=1).alignment = Alignment(indent=1, vertical="center")
            ws.cell(row=r, column=1).fill = _fill(bg_row)

            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            pc = ws.cell(row=r, column=2, value=_mx2(peer_v))
            pc.font      = Font(size=9, color=C_DARK)
            pc.fill      = _fill(bg_row)
            pc.alignment = Alignment(horizontal="center", vertical="center")

            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
            tc = ws.cell(row=r, column=4, value=_mx2(tgt_v))
            tc.font      = Font(size=9, color=C_DARK)
            tc.fill      = _fill(bg_row)
            tc.alignment = Alignment(horizontal="center", vertical="center")

            # vs-target delta
            delta = None
            delta_str = "—"
            if peer_v is not None and tgt_v is not None:
                delta = peer_v - tgt_v
                delta_str = f"{delta:+.1f}x"
            ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=NC)
            dc = ws.cell(row=r, column=6, value=delta_str)
            dc.font = Font(size=9,
                           color=("375623" if delta and delta > 0
                                  else "C00000" if delta and delta < 0
                                  else C_DARK))
            dc.fill      = _fill(bg_row)
            dc.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[r].height = 14
            r += 1

        # Implied VPS range vs DCF
        if rng and len(rng) >= 3:
            lo, mid_v, hi = rng[0], rng[1], rng[2]
            rng_str = f"{ccy} {lo:.2f}  —  {mid_v:.2f}  —  {hi:.2f}"
            dcf_str = f"DCF base: {ccy} {vps:.2f}" if vps else ""
            _wrap_row(r, 1, NC,
                      f"Peer-implied VPS range (low / mid / high):  {rng_str}     {dcf_str}",
                      size=9, bold=True, bg=C_AMBER, fg=C_DARK, min_h=15)
            r += 1

    _spacer(r); r += 1

    # ======================================================================
    # Disclaimer
    # ======================================================================
    _full(r,
          "v0.1  |  FOR REVIEW ONLY  |  Generated automatically from public data.  "
          "Not investment advice.  All assumptions require qualified analyst review.",
          bg="F2F2F2", fg="999999", bold=False, italic=True, size=8, height=16)


# ---------------------------------------------------------------------------
# Tab 2 — Historical Financials
# ---------------------------------------------------------------------------

def _build_historical_sheet(wb: Workbook, valued: dict, meta: dict):
    ws = wb.create_sheet("Historical")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B4"

    years_data = valued.get("canonical_by_year", {})
    stats      = valued.get("stats", {})
    ccy        = stats.get("currency", "")
    years      = sorted(years_data.keys())   # oldest → newest

    _col_width(ws, 1, 30)
    for c in range(2, len(years) + 2):
        _col_width(ws, c, 12)

    r = 1
    _merge_write(ws, r, 1, len(years) + 1,
                 f"Historical Financials  —  {meta['company']} ({meta['ticker']})  |  ({ccy} millions)",
                 font=Font(bold=True, size=12, color=C_WHITE),
                 fill=_fill(C_NAVY), height=22,
                 align=Alignment(horizontal="left", vertical="center", indent=2))

    r += 1
    _label_cell(ws, r, 1, "")
    for i, yr in enumerate(years):
        _header_cell(ws, r, 2 + i, f"{yr}A", bg=C_HIST)

    # Helper to write a data section.
    # rows_def items: (label, field, fmt, bold) — or with an optional 5th element
    # which is a denominator_field; if present, writes field/denominator_field as a ratio
    # (skips _m() scaling — used for margin rows).
    def _data_section(title, rows_def):
        nonlocal r
        r += 1
        _section_header(ws, r, 1, len(years) + 1, title, color=C_NAVY)
        for row in rows_def:
            label, field, fmt, bold = row[:4]
            denom_field = row[4] if len(row) > 4 else None
            r += 1
            _label_cell(ws, r, 1, label, bold=bold, indent=0 if bold else 1)
            for i, yr in enumerate(years):
                if denom_field:
                    num   = years_data[yr].get(field)
                    denom = years_data[yr].get(denom_field)
                    v = (num / denom) if (num is not None and denom) else None
                else:
                    raw = years_data[yr].get(field)
                    v   = _m(raw) if raw is not None else None
                cell = ws.cell(row=r, column=2 + i, value=v)
                cell.number_format = fmt
                cell.font = Font(bold=bold, size=10, color=C_DARK)
                cell.alignment = Alignment(horizontal="right")
                cell.fill = _fill(C_LGRAY if (2 + i) % 2 == 0 else C_WHITE)

    _data_section("Income Statement", [
        ("Revenue",          "revenue",          FMT_M, True),
        ("  Gross Profit",   "gross_profit",      FMT_M, False),
        ("  Gross Margin",   "gross_profit",      "0.0%",    False, "revenue"),
        ("EBIT",             "ebit",              FMT_M, True),
        ("  EBIT Margin",    "ebit",              "0.0%",    False, "revenue"),
        ("EBITDA",           "ebitda",            FMT_M, False),
        ("  D&A",            "da",                FMT_M, False),
        ("Pre-Tax Income",   "pre_tax_income",    FMT_M, False),
        ("  Tax Provision",  "tax_provision",     FMT_M, False),
        ("Net Income",       "net_income",        FMT_M, True),
        ("  Interest Expense","interest_expense", FMT_M, False),
    ])

    _data_section("Cash Flow", [
        ("Operating Cash Flow", "operating_cash_flow", FMT_M, True),
        ("  Capital Expenditure","capex",               FMT_M, False),
        ("Free Cash Flow",       "free_cash_flow",      FMT_M, True),
        ("  D&A",                "da",                  FMT_M, False),
        ("  Δ Working Capital",  "change_in_working_cap",FMT_M,False),
    ])

    _data_section("Balance Sheet", [
        ("Cash & Equivalents",  "cash",               FMT_M, True),
        ("Total Debt",          "debt",               FMT_M, False),
        ("  Long-term Debt",    "long_term_debt",     FMT_M, False),
        ("  Lease Liabilities", "lease_liabilities",  FMT_M, False),
        ("Total Assets",        "total_assets",       FMT_M, True),
        ("Total Equity",        "total_equity",       FMT_M, False),
        ("  Accounts Receivable","accounts_receivable",FMT_M,False),
        ("  Inventory",         "inventory",          FMT_M, False),
        ("  Accounts Payable",  "accounts_payable",   FMT_M, False),
        ("Shares Outstanding",  "shares_outstanding", FMT_M, False),
    ])

    # Ratios section — computed inline
    r += 1
    _section_header(ws, r, 1, len(years) + 1, "Key Ratios", color=C_BLUE)
    ratio_defs = [
        ("EPS (diluted)", "eps_diluted", "0.00"),
        ("Net Profit Margin", "net_profit_margin_pct", "0.0%"),
    ]
    for label, field, fmt in ratio_defs:
        r += 1
        _label_cell(ws, r, 1, label)
        for i, yr in enumerate(years):
            v = years_data[yr].get(field)
            cell = ws.cell(row=r, column=2 + i, value=v)
            cell.number_format = fmt
            cell.font = Font(size=10, color=C_DARK)
            cell.alignment = Alignment(horizontal="right")


# ---------------------------------------------------------------------------
# Tab 3 — Forecast (3 scenarios)
# ---------------------------------------------------------------------------

def _build_forecast_sheet(wb: Workbook, valued: dict, meta: dict):
    ws = wb.create_sheet("Forecast")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B4"

    forecast = valued.get("forecast", {})
    stats    = valued.get("stats", {})
    ccy      = stats.get("currency", "")
    n_years  = valued.get("forecast_years", 5)

    _col_width(ws, 1, 28)
    for c in range(2, n_years + 2):
        _col_width(ws, c, 12)

    r = 1
    _merge_write(ws, r, 1, n_years + 1,
                 f"Forecast Model  —  {meta['company']} ({meta['ticker']})  |  ({ccy} millions)",
                 font=Font(bold=True, size=12, color=C_WHITE),
                 fill=_fill(C_NAVY), height=22,
                 align=Alignment(horizontal="left", vertical="center", indent=2))

    scenario_colors = {"base": C_HIST, "bull": C_GREEN, "bear": C_RED}
    fcf_fields = [
        ("Revenue",          "revenue",        FMT_M, True),
        ("  Revenue Growth", "revenue_growth", "0.0%",    False),
        ("EBIT Margin",      "ebit_margin",    "0.0%",    False),
        ("EBIT",             "ebit",           FMT_M, True),
        ("  Tax",            "tax",            FMT_M, False),
        ("NOPAT",            "nopat",          FMT_M, True),
        ("  D&A",            "da",             FMT_M, False),
        ("  Capex",          "capex",          FMT_M, False),
        ("  Δ NWC",          "delta_nwc",      FMT_M, False),
        ("FCFF",             "fcff",           FMT_M, True),
        ("  FCFF Margin",    "fcff_margin",    "0.0%",    False),
        ("  Cash Conversion","cash_conversion","0.00",    False),
    ]

    for sc_name in ["base", "bull", "bear"]:
        proj = forecast.get(sc_name, {})
        if not proj:
            continue
        year_keys = sorted(proj.keys())
        bg = scenario_colors.get(sc_name, C_LGRAY)

        r += 1
        _section_header(ws, r, 1, n_years + 1,
                        f"{sc_name.upper()} CASE", color=C_NAVY)
        r += 1
        _label_cell(ws, r, 1, "")
        for i, yk in enumerate(year_keys):
            cal = proj[yk]["calendar_year"]
            _header_cell(ws, r, 2 + i, f"{cal}E", bg=bg)

        _PCT_FMTS = {"0.0%", "0.00"}  # fields that are already ratios/per-share
        for label, field, fmt, bold in fcf_fields:
            r += 1
            _label_cell(ws, r, 1, label, bold=bold,
                        indent=0 if bold else 1)
            for i, yk in enumerate(year_keys):
                raw = proj[yk].get(field)
                # Apply _m() to monetary fields; leave ratios/margins as-is
                v = raw if fmt in _PCT_FMTS else _m(raw)
                cell = ws.cell(row=r, column=2 + i, value=v)
                cell.number_format = fmt
                cell.font = Font(bold=bold, size=10, color=C_DARK)
                cell.alignment = Alignment(horizontal="right")
                cell.fill = _fill(bg)


# ---------------------------------------------------------------------------
# Tab 4 — Sensitivity
# ---------------------------------------------------------------------------

def _build_sensitivity_sheet(wb: Workbook, valued: dict, meta: dict):
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_view.showGridLines = False

    sens         = valued.get("sensitivity", {})
    stats        = valued.get("stats", {})
    current_price = stats.get("current_price")
    wacc_vals    = sens.get("wacc_values", [])
    growth_vals  = sens.get("growth_values", [])
    table        = sens.get("table", {})

    _col_width(ws, 1, 18)
    for c in range(2, len(wacc_vals) + 2):
        _col_width(ws, c, 13)

    r = 1
    _merge_write(ws, r, 1, len(wacc_vals) + 1,
                 f"Sensitivity Analysis  —  Value per Share ({stats.get('currency', '')})  |  Base case FCFs",
                 font=Font(bold=True, size=12, color=C_WHITE),
                 fill=_fill(C_NAVY), height=22,
                 align=Alignment(horizontal="left", vertical="center", indent=2))

    r += 1
    ccy_s = stats.get("currency", "")
    price_str = f"{ccy_s} {current_price:.2f}" if current_price is not None else "n/a"
    note = (f"Current price: {price_str}  |  "
            f"Green = above current price  |  Red = below current price  |  "
            f"Rows = terminal growth  |  Columns = WACC")
    _merge_write(ws, r, 1, len(wacc_vals) + 1, note,
                 font=Font(italic=True, size=9, color="595959"),
                 fill=_fill("F2F2F2"), height=14,
                 align=Alignment(horizontal="left", vertical="center", indent=2))

    r += 2
    # Header row
    _header_cell(ws, r, 1, "g  ↓  /  WACC  →", bg=C_NAVY)
    for i, w in enumerate(wacc_vals):
        _header_cell(ws, r, 2 + i, w, bg=C_BLUE)

    for g in growth_vals:
        r += 1
        _header_cell(ws, r, 1, g, bg=C_LGRAY)
        for i, w in enumerate(wacc_vals):
            vps = table.get(g, {}).get(w)
            cell = ws.cell(row=r, column=2 + i, value=vps)
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if vps is None:
                cell.value = "n/a"
                cell.font = Font(size=10, color="999999")
            else:
                above = current_price and vps > current_price
                below = current_price and vps < current_price
                cell.fill = _fill(C_GREEN if above else C_RED if below else C_LGRAY)
                cell.font = Font(bold=(above or below), size=10,
                                 color="375623" if above else "C00000" if below else C_DARK)
            ws.row_dimensions[r].height = 18


# ---------------------------------------------------------------------------
# Tab 6 — WACC Calculation
# ---------------------------------------------------------------------------

def _build_wacc_sheet(wb: Workbook, valued: dict, meta: dict):
    ws = wb.create_sheet("WACC")
    ws.sheet_view.showGridLines = False

    # Columns: A(label, 34) | B(gap, 3) | C(value col, 18)
    _col_width(ws, 1, 34)
    _col_width(ws, 2, 3)
    _col_width(ws, 3, 16)  # data col — uniform with year headers below
    _col_width(ws, 4, 16)  # second year col

    wc      = valued.get("wacc_components", {})
    stats   = valued.get("stats", {})
    norm    = valued.get("normalised", {})
    years_d = valued.get("canonical_by_year", {})
    years   = sorted(years_d.keys(), reverse=True)

    ccy     = stats.get("currency", "")
    company = meta["company"]
    ticker  = meta["ticker"]
    tax     = wc.get("tax_rate")

    # --- helpers local to this sheet ---
    def _sh(r, c1, c2, text, color=C_NAVY):
        _section_header(ws, r, c1, c2, text, color)

    def _lbl(r, text, bold=False, indent=1):
        _label_cell(ws, r, 1, text, bold=bold, indent=indent)

    def _val(r, value, fmt="0.0%", bold=False, bg=None):
        _calc_cell(ws, r, 3, value, fmt=fmt, bold=bold, bg=bg,
                   color=C_INPUT if not bold else C_DARK)

    def _row_lv(r, label, value, fmt="0.0%", bold=False, bg=None,
                label_indent=1):
        _lbl(r, label, bold=bold, indent=label_indent)
        _val(r, value, fmt=fmt, bold=bold, bg=bg)
        if bg:
            ws.cell(row=r, column=1).fill = _fill(bg)

    r = 1
    # Title
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(row=r, column=1,
                value=f"WACC Calculation  —  {company} ({ticker})")
    c.font = Font(bold=True, size=13, color=C_NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 22
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    sub = ws.cell(row=r, column=1,
                  value=f"All figures in {ccy}. Source: Yahoo Finance / country defaults.")
    sub.font = Font(italic=True, size=9, color="808080")
    ws.row_dimensions[r].height = 13
    r += 2

    # ── Section 1: Tax Rate ──────────────────────────────────────────────────
    _sh(r, 1, 3, "Tax Rate")
    r += 1
    _row_lv(r, "Effective Tax Rate", tax, "0.0%")
    r += 2

    # ── Section 2: Cost of Debt ──────────────────────────────────────────────
    _sh(r, 1, 3, "Cost of Debt")
    r += 1

    # Pull raw data for the two most recent years to show the working
    yr_labels = years[:2]
    debt_vals  = [years_d[y].get("debt")             for y in yr_labels]
    ie_vals    = [years_d[y].get("interest_expense")  for y in yr_labels]

    # Year column headers on their own row (cannot write into the merged section header row)
    for i, yr in enumerate(yr_labels):
        hdr_col = 3 + i
        hc = ws.cell(row=r, column=hdr_col, value=yr)
        hc.font      = Font(bold=True, size=10, color=C_WHITE)
        hc.fill      = _fill(C_NAVY)
        hc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 16
    r += 1

    def _debt_row(r, label, vals, fmt=FMT_M, bold=False):
        _lbl(r, label, bold=bold)
        for i, v in enumerate(vals):
            disp, f = _smart_m(v)
            ws.cell(row=r, column=3 + i, value=disp).number_format = fmt
            ws.cell(row=r, column=3 + i).font = Font(size=10, color=C_DARK, bold=bold)
            ws.cell(row=r, column=3 + i).alignment = Alignment(horizontal="right")
        ws.row_dimensions[r].height = 15

    _debt_row(r, "Total Debt", debt_vals)
    r += 1

    # Average debt (only needs one cell — most recent pair)
    avg_debt = None
    if debt_vals[0] and (len(debt_vals) > 1 and debt_vals[1]):
        avg_debt = (debt_vals[0] + debt_vals[1]) / 2
    avg_disp, avg_fmt = _smart_m(avg_debt)
    _lbl(r, "Average Debt")
    ws.cell(row=r, column=3, value=avg_disp).number_format = avg_fmt
    ws.cell(row=r, column=3).font = Font(size=10, color=C_DARK)
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="right")
    ws.row_dimensions[r].height = 15
    r += 1

    _debt_row(r, "Interest Expense (net)", ie_vals)
    r += 1

    cod_pretax = wc.get("cost_of_debt_pretax")
    cod_method = wc.get("cost_of_debt_method", "")
    _lbl(r, f"Pre-Tax Cost of Debt  [{cod_method}]", bold=True)
    _calc_cell(ws, r, 3, cod_pretax, fmt="0.0%", bold=True, bg=C_LGRAY)
    ws.cell(row=r, column=1).fill = _fill(C_LGRAY)
    r += 1

    cod_aftertax = wc.get("cost_of_debt_aftertax")
    _lbl(r, "Post-Tax Cost of Debt  [pre-tax × (1 − tax rate)]")
    _calc_cell(ws, r, 3, cod_aftertax, fmt="0.0%")
    r += 2

    # ── Section 3: Cost of Equity ────────────────────────────────────────────
    _sh(r, 1, 3, "Cost of Equity  (CAPM with Blume-Adjusted Beta)")
    r += 1

    rf          = wc.get("risk_free_rate")
    erp         = wc.get("equity_risk_premium")
    beta_raw    = wc.get("beta_raw")
    beta_adj    = wc.get("beta_adjusted")
    ke          = wc.get("cost_of_equity")
    country     = wc.get("country", "")

    _row_lv(r, f"Risk-Free Rate  ({country} 10yr gov't bond, country default)", rf)
    r += 1
    _row_lv(r, "Equity Risk Premium  (Damodaran country estimate)", erp)
    r += 1
    _row_lv(r, "Beta (raw — Yahoo Finance)", beta_raw, fmt="0.00")
    r += 1
    _lbl(r, "Adjusted Beta  [= (²⁄₃ × raw) + (¹⁄₃)]  Blume adjustment")
    _calc_cell(ws, r, 3, beta_adj, fmt="0.00", bold=True, bg=C_LGRAY)
    ws.cell(row=r, column=1).fill = _fill(C_LGRAY)
    r += 1
    _lbl(r, "Cost of Equity  [Rf + (Adj β × ERP)]", bold=True)
    _calc_cell(ws, r, 3, ke, fmt="0.0%", bold=True, bg=C_LGRAY)
    ws.cell(row=r, column=1).fill = _fill(C_LGRAY)
    r += 2

    # ── Section 4: Weightings ────────────────────────────────────────────────
    _sh(r, 1, 3, "Capital Structure Weightings  (Market Value)")
    r += 1

    market_cap  = wc.get("market_cap_used") or stats.get("market_cap")
    debt_bridge = wc.get("debt_used")
    w_equity    = wc.get("weight_equity")
    w_debt      = wc.get("weight_debt")
    wt_method   = wc.get("weight_method", "")

    mc_disp, mc_fmt = _smart_m(market_cap)
    db_disp, db_fmt = _smart_m(debt_bridge)
    total_disp, total_fmt = _smart_m(
        (market_cap or 0) + (debt_bridge or 0)
    )

    _lbl(r, "Market Capitalisation  (equity weight basis)")
    _calc_cell(ws, r, 3, mc_disp, fmt=mc_fmt)
    r += 1
    _lbl(r, "Total Debt  (balance sheet)")
    _calc_cell(ws, r, 3, db_disp, fmt=db_fmt)
    r += 1
    _lbl(r, "Total Value  (market cap + debt)", bold=True)
    _calc_cell(ws, r, 3, total_disp, fmt=total_fmt, bold=True)
    r += 1
    _lbl(r, f"Weight of Equity  [{wt_method}]", bold=True)
    _calc_cell(ws, r, 3, w_equity, fmt="0.0%", bold=True, bg=C_LGRAY)
    ws.cell(row=r, column=1).fill = _fill(C_LGRAY)
    r += 1
    _lbl(r, "Weight of Debt", bold=True)
    _calc_cell(ws, r, 3, w_debt, fmt="0.0%", bold=True, bg=C_LGRAY)
    ws.cell(row=r, column=1).fill = _fill(C_LGRAY)
    r += 2

    # ── Section 5: WACC Summary ──────────────────────────────────────────────
    _sh(r, 1, 3, "Weighted Average Cost of Capital (WACC)")
    r += 1

    wacc_val = valued.get("assumptions", {}).get("wacc", {}).get("value")

    rows_wacc = [
        ("Weight of Equity",        w_equity,     "0.0%"),
        ("Cost of Equity (Ke)",      ke,           "0.0%"),
        ("Weight of Debt",           w_debt,       "0.0%"),
        ("Pre-Tax Cost of Debt",     cod_pretax,   "0.0%"),
        ("Tax Rate",                 tax,          "0.0%"),
        ("Post-Tax Cost of Debt",    cod_aftertax, "0.0%"),
    ]
    for label, val, fmt in rows_wacc:
        _lbl(r, label)
        _calc_cell(ws, r, 3, val, fmt=fmt)
        r += 1

    ws.row_dimensions[r].height = 3
    r += 1
    _lbl(r, "WACC  [= Ke×We + Kd(1−t)×Wd]", bold=True)
    _calc_cell(ws, r, 3, wacc_val, fmt="0.00%", bold=True,
               bg=C_NAVY, color=C_WHITE)
    ws.cell(row=r, column=1).fill = _fill(C_NAVY)
    ws.cell(row=r, column=1).font = Font(bold=True, size=11, color=C_WHITE)
    ws.row_dimensions[r].height = 20
    r += 2

    # ── Footnotes ────────────────────────────────────────────────────────────
    notes = [
        "Blume adjustment: adjusted β = (²⁄₃ × raw β) + ¹⁄₃. "
        "Reflects empirical mean-reversion of betas over time.",
        "Risk-free rate and ERP are country defaults (early 2026 estimates). "
        "Analyst must confirm against current market rates before use.",
        "Cost of debt derived from interest expense ÷ average debt (last 2 years). "
        "Falls back to sector default if data unavailable.",
    ]
    for note in notes:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        nc = ws.cell(row=r, column=1, value=f"• {note}")
        nc.font      = Font(italic=True, size=8, color="808080")
        nc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 26
        r += 1


# ---------------------------------------------------------------------------
# Tab 7 — Comparables
# ---------------------------------------------------------------------------

def _build_comparables_sheet(wb: Workbook, valued: dict, meta: dict):
    comps = valued.get("comparables")
    ws = wb.create_sheet("Comparables")
    ws.sheet_view.showGridLines = False

    stats  = valued.get("stats", {})
    ccy    = stats.get("currency", "")
    sector = stats.get("sector", "")

    _col_width(ws, 1, 28)   # company name
    _col_width(ws, 2, 10)   # ticker
    _col_width(ws, 3, 14)   # EV
    _col_width(ws, 4, 14)   # Revenue
    _col_width(ws, 5, 14)   # EBITDA
    _col_width(ws, 6, 14)   # EBIT
    _col_width(ws, 7, 11)   # EV/Rev
    _col_width(ws, 8, 11)   # EV/EBITDA
    _col_width(ws, 9, 11)   # EV/EBIT
    _col_width(ws, 10, 9)   # P/E

    r = 1
    # Title
    _merge_write(ws, r, 1, 10,
                 f"Comparable Companies  —  {meta['company']} ({meta['ticker']})  "
                 f"|  Sector: {sector}",
                 font=Font(bold=True, size=12, color=C_WHITE),
                 fill=_fill(C_NAVY), height=22,
                 align=Alignment(horizontal="left", vertical="center", indent=2))

    # No comparables data — write a placeholder and return
    if not comps or not comps.get("available"):
        r += 2
        reason = (comps or {}).get("reason", "Comparables not fetched.")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        c = ws.cell(row=r, column=1, value=f"No comparables data available. {reason}")
        c.font = Font(italic=True, size=10, color="808080")
        return

    # Notes row
    r += 1
    notes_str = "  |  ".join((comps.get("notes") or []))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    nc = ws.cell(row=r, column=1, value=notes_str or "")
    nc.font = Font(italic=True, size=9, color="595959")
    nc.fill = _fill("F2F2F2")
    ws.row_dimensions[r].height = 13

    r += 2
    # Column headers
    headers = ["Company", "Ticker", "EV", "Revenue", "EBITDA", "EBIT",
               "EV/Rev", "EV/EBITDA", "EV/EBIT", "P/E"]
    ws.row_dimensions[r].height = 16
    for i, h in enumerate(headers, 1):
        _header_cell(ws, r, i, h, bg=C_NAVY)

    def _company_row(d: dict, bg_name: str, bg_num: str, is_target: bool = False):
        nonlocal r
        r += 1
        ws.row_dimensions[r].height = 15
        bold = is_target

        # Company name
        c = ws.cell(row=r, column=1, value=d.get("name", ""))
        c.font = Font(bold=bold, size=10, color=C_DARK)
        c.fill = _fill(bg_name)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Ticker
        t = ws.cell(row=r, column=2, value=d.get("ticker", ""))
        t.font = Font(bold=bold, size=10, color=C_INPUT)
        t.fill = _fill(bg_name)
        t.alignment = Alignment(horizontal="center", vertical="center")

        # Financial values (auto B/M scale)
        for col, field in [(3, "ev"), (4, "revenue"), (5, "ebitda"), (6, "ebit")]:
            val = d.get(field)
            disp_v, fmt = _smart_m(val)
            cell = ws.cell(row=r, column=col, value=disp_v)
            cell.number_format = fmt
            cell.font = Font(bold=bold, size=10, color=C_DARK)
            cell.fill = _fill(bg_num)
            cell.alignment = Alignment(horizontal="right", vertical="center")

        # Multiples
        for col, field, fmt in [
            (7,  "ev_revenue",  "0.0x"),
            (8,  "ev_ebitda",   "0.0x"),
            (9,  "ev_ebit",     "0.0x"),
            (10, "pe",          "0.0x"),
        ]:
            val = d.get(field)
            cell = ws.cell(row=r, column=col, value=val)
            cell.number_format = fmt
            cell.font = Font(bold=bold, size=10, color=C_DARK)
            cell.fill = _fill(bg_num)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if val is not None:
                cell.border = _top_border() if is_target else None

    # Target company row (yellow highlight)
    target = comps.get("target", {})
    _company_row(target, "FFFF00", "FFFACD", is_target=True)

    # Spacer
    r += 1
    ws.row_dimensions[r].height = 4

    # Peer rows
    for i, peer in enumerate(comps.get("peers", [])):
        bg_n = C_HIST if i % 2 == 0 else C_WHITE
        _company_row(peer, bg_n, bg_n)

    # Median row
    med = comps.get("median", {})
    if med:
        r += 1
        ws.row_dimensions[r].height = 15
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        mc = ws.cell(row=r, column=1, value="Peer Median")
        mc.font = Font(bold=True, size=10, color=C_WHITE)
        mc.fill = _fill(C_BLUE)

        for col, field, fmt in [
            (7,  "ev_revenue", "0.0x"),
            (8,  "ev_ebitda",  "0.0x"),
            (9,  "ev_ebit",    "0.0x"),
            (10, "pe",         "0.0x"),
        ]:
            val = med.get(field)
            cell = ws.cell(row=r, column=col, value=val)
            cell.number_format = fmt
            cell.font = Font(bold=True, size=10, color=C_WHITE)
            cell.fill = _fill(C_BLUE)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = _top_border()
        # Blank financial cols
        for col in [3, 4, 5, 6]:
            ws.cell(row=r, column=col).fill = _fill(C_BLUE)

    # ── Implied valuation section ────────────────────────────────────────────
    impl = comps.get("implied_from_peers", {})
    if impl:
        r += 3
        _section_header(ws, r, 1, 10,
                        f"Implied Valuation from Peer Multiples  |  {ccy}",
                        color=C_NAVY)
        r += 1

        # Header
        ws.row_dimensions[r].height = 14
        for col, label in [(1, "Multiple Used"), (2, "Implied EV"),
                           (3, "Implied VPS"), (4, "vs DCF VPS")]:
            _header_cell(ws, r, col, label, bg=C_BLUE)

        dcf_vps = target.get("dcf_vps")
        bridge  = (valued.get("equity_bridge_inputs") or {})
        shares  = bridge.get("shares_outstanding")

        rows_impl = [
            ("EV / Revenue",
             impl.get("ev_from_revenue_multiple"),
             impl.get("vps_from_revenue_multiple")),
            ("EV / EBITDA",
             impl.get("ev_from_ebitda_multiple"),
             impl.get("vps_from_ebitda_multiple")),
            ("EV / EBIT",
             impl.get("ev_from_ebit_multiple"),
             impl.get("vps_from_ebit_multiple")),
        ]

        for label, impl_ev, impl_vps in rows_impl:
            r += 1
            ws.row_dimensions[r].height = 15
            ws.cell(row=r, column=1, value=label).font = Font(size=10)
            ev_d, ev_f = _smart_m(impl_ev)
            c2 = ws.cell(row=r, column=2, value=ev_d)
            c2.number_format = ev_f
            c2.font = Font(size=10)
            c2.alignment = Alignment(horizontal="right")

            c3 = ws.cell(row=r, column=3, value=impl_vps)
            c3.number_format = "0.00"
            c3.font = Font(bold=True, size=10)
            c3.alignment = Alignment(horizontal="right")
            c3.fill = _fill(C_FORE)

            if dcf_vps and impl_vps:
                diff = impl_vps - dcf_vps
                is_p = diff > 0
                c4 = ws.cell(row=r, column=4, value=diff)
                c4.number_format = "0.00"
                c4.font = Font(size=10,
                               color="375623" if is_p else "C00000")
                c4.alignment = Alignment(horizontal="right")

        # VPS range summary
        rng = impl.get("vps_range", [])
        if len(rng) == 3:
            r += 2
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            summary = (f"Peer-implied VPS range:  "
                       f"{ccy} {rng[0]:.2f}  (low)  —  "
                       f"{ccy} {rng[1]:.2f}  (median)  —  "
                       f"{ccy} {rng[2]:.2f}  (high)")
            sc = ws.cell(row=r, column=1, value=summary)
            sc.font = Font(bold=True, size=10, color=C_NAVY)
            sc.fill = _fill(C_HIST)
            sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[r].height = 16

        # DCF comparison note
        if dcf_vps:
            r += 1
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            note = (f"DCF intrinsic value: {ccy} {dcf_vps:.2f}  |  "
                    f"Current price: {ccy} {target.get('price', 0):.2f}")
            nc = ws.cell(row=r, column=1, value=note)
            nc.font = Font(italic=True, size=9, color="595959")
            ws.row_dimensions[r].height = 13

    # ── Peer Selection Scores ─────────────────────────────────────────────────
    selection = (comps.get("peer_selection") or {})
    scored_peers = selection.get("selected") or comps.get("peers", [])

    if scored_peers and any(p.get("score") is not None for p in scored_peers):
        r += 3
        _col_width(ws, 1, 28)
        _col_width(ws, 2, 8)
        _col_width(ws, 3, 10)
        _col_width(ws, 4, 12)
        _col_width(ws, 5, 12)
        _col_width(ws, 6, 12)
        _col_width(ws, 7, 50)
        _section_header(ws, r, 1, 7, "Peer Selection Scores", color=C_NAVY)
        r += 1

        for col, txt in [
            (1, "Company"), (2, "Ticker"), (3, "Score /100"),
            (4, "Industry"), (5, "Size"), (6, "Financials"),
            (7, "Selection Notes"),
        ]:
            hc = ws.cell(row=r, column=col, value=txt)
            hc.font      = Font(bold=True, size=9, color=C_WHITE)
            hc.fill      = _fill(C_BLUE)
            hc.alignment = Alignment(horizontal="center" if col > 1 else "left",
                                     vertical="center", indent=1 if col == 1 else 0)
        ws.row_dimensions[r].height = 14
        r += 1

        for p in scored_peers:
            sc      = p.get("score")
            bkdn    = p.get("score_breakdown", {})
            notes_p = "  |  ".join(p.get("match_notes", []))
            sc_bg   = (C_GREEN if (sc or 0) >= 75 else
                       C_AMBER if (sc or 0) >= 50 else C_RED)

            for col, val, fmt, align in [
                (1, p.get("name", ""),                    None,   "left"),
                (2, p.get("ticker", ""),                  None,   "center"),
                (3, sc,                                   "0.0",  "center"),
                (4, bkdn.get("industry"),                 "0.0",  "center"),
                (5, bkdn.get("size"),                     "0.0",  "center"),
                (6, bkdn.get("financial"),                "0.0",  "center"),
                (7, notes_p,                              None,   "left"),
            ]:
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = Font(
                    bold=(col == 3),
                    size=9,
                    color=(C_DARK if col != 3 else
                           "375623" if (sc or 0) >= 75 else
                           "7F6000" if (sc or 0) >= 50 else "C00000"),
                )
                cell.fill = _fill(sc_bg if col == 3 else
                                  (C_LGRAY if r % 2 == 0 else C_WHITE))
                cell.alignment = Alignment(
                    horizontal=align, vertical="center",
                    indent=1 if col in (1, 7) else 0,
                    wrap_text=(col == 7),
                )
                if fmt:
                    cell.number_format = fmt
            ws.row_dimensions[r].height = max(14, min(40, 9 * (len(notes_p) // 80 + 1) + 4))
            r += 1

    # ── Rejected / Excluded Peers ─────────────────────────────────────────────
    rejected = selection.get("rejected", [])
    if rejected:
        r += 2
        _section_header(ws, r, 1, 7,
                        f"Rejected / Excluded Peers  ({len(rejected)} candidates)",
                        color=C_BLUE)
        r += 1
        sub = ws.cell(row=r, column=1,
                      value="Candidates considered but not selected. "
                            "Scores shown where full evaluation was completed.")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        sub.font = Font(size=9, italic=True, color="595959")
        sub.alignment = Alignment(indent=2)
        ws.row_dimensions[r].height = 13
        r += 1

        for col, txt in [
            (1, "Company"), (2, "Ticker"), (3, "Score /100"), (4, "Rejection Reason"),
        ]:
            hc = ws.cell(row=r, column=col, value=txt)
            hc.font      = Font(bold=True, size=9, color=C_WHITE)
            hc.fill      = _fill(C_BLUE)
            hc.alignment = Alignment(horizontal="center" if col in (2, 3) else "left",
                                     vertical="center", indent=1 if col in (1, 4) else 0)
        ws.row_dimensions[r].height = 14
        r += 1

        for rej in rejected:
            sc_r   = rej.get("score")
            reason = rej.get("rejection_reason", "—")
            for col, val, align in [
                (1, rej.get("name", rej.get("ticker", "—")), "left"),
                (2, rej.get("ticker", "—"),                  "center"),
                (3, f"{sc_r:.0f}" if sc_r is not None else "—", "center"),
                (4, reason,                                  "left"),
            ]:
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = Font(size=9, color=C_DARK)
                cell.fill = _fill(C_LGRAY if r % 2 == 0 else C_WHITE)
                cell.alignment = Alignment(
                    horizontal=align, vertical="center",
                    indent=1 if col in (1, 4) else 0,
                    wrap_text=(col == 4),
                )
            ws.row_dimensions[r].height = max(14, min(36, 9 * (len(reason) // 60 + 1) + 4))
            r += 1


# ---------------------------------------------------------------------------
# Tab 8 — Pipeline / Data Quality
# ---------------------------------------------------------------------------

def _build_raw_data_sheet(wb: Workbook, valued: dict, meta: dict):
    """Tab: Raw reconciled data with per-cell source attribution (colour-coded)."""
    ws = wb.create_sheet("Raw Data")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B5"

    reconciled = valued.get("reconciled", {})
    stats      = valued.get("stats", {})
    ccy        = stats.get("currency", "")

    if not reconciled:
        ws.cell(row=1, column=1, value="No reconciled data available.")
        return

    years = sorted(reconciled.keys())   # oldest → newest (left to right)
    n     = len(years)

    # Source → cell background colour
    _SRC_BG = {
        "yahoo":       "D9E1F2",   # soft blue
        "edgar":       "E2EFDA",   # soft green
        "macrotrends": "FFF2CC",   # soft amber
    }

    _col_width(ws, 1, 30)
    for c in range(2, 2 + n):
        _col_width(ws, c, 13)

    r = 1
    _merge_write(ws, r, 1, 1 + n,
                 f"Raw Scraped Data  —  {meta['company']} ({meta['ticker']})",
                 font=Font(bold=True, size=12, color=C_WHITE),
                 fill=_fill(C_NAVY), height=22,
                 align=Alignment(horizontal="left", vertical="center", indent=2))
    r += 1

    # Legend
    lc = ws.cell(row=r, column=1,
                 value="Source key:   Yahoo Finance (blue)   |   SEC EDGAR (green)   |   Macrotrends (amber)")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1 + n)
    lc.font      = Font(size=9, italic=True, color="595959")
    lc.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[r].height = 14
    r += 2

    # Year header row
    _label_cell(ws, r, 1, f"Metric  ({ccy}M unless noted)", bold=True)
    for i, yr in enumerate(years):
        hc = ws.cell(row=r, column=2 + i, value=str(yr))
        hc.font      = Font(bold=True, size=10, color=C_WHITE)
        hc.fill      = _fill(C_BLUE)
        hc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 16
    r += 1

    # (label, field, scale_to_m, number_format)
    metrics_sections = [
        ("Income Statement", [
            ("Revenue",           "revenue",           True,  FMT_M),
            ("Gross Profit",      "gross_profit",      True,  FMT_M),
            ("Operating Income",  "operating_income",  True,  FMT_M),
            ("EBITDA",            "ebitda",            True,  FMT_M),
            ("Net Income",        "net_income",        True,  FMT_M),
        ]),
        ("Cash Flow", [
            ("Operating Cash Flow", "operating_cash_flow", True, FMT_M),
            ("Capital Expenditure", "capex",               True, FMT_M),
            ("Free Cash Flow",      "free_cash_flow",      True, FMT_M),
        ]),
        ("Balance Sheet", [
            ("Total Assets",        "total_assets",         True,  FMT_M),
            ("Total Debt",          "total_debt",           True,  FMT_M),
            ("Cash & Equivalents",  "cash_and_equivalents", True,  FMT_M),
        ]),
        ("Per Share & Margins", [
            ("EPS (diluted)",         "eps_diluted",            False, "0.00"),
            ("Net Profit Margin %",   "net_profit_margin_pct",  False, "0.0%"),
            ("Shares Outstanding",    "shares_outstanding",     True,  FMT_M),
        ]),
    ]

    for section_title, rows in metrics_sections:
        _section_header(ws, r, 1, 1 + n, section_title, color=C_NAVY)
        r += 1
        for label, field, scale, fmt in rows:
            _label_cell(ws, r, 1, label)
            for i, yr in enumerate(years):
                yr_data  = reconciled.get(yr, {})
                src_data = yr_data.get("_sources", {})
                src      = src_data.get(field, "yahoo")
                bg       = _SRC_BG.get(src, C_WHITE)
                raw      = yr_data.get(field)

                if raw is None:
                    v = None
                elif scale:
                    v = _m(raw)
                else:
                    try:
                        v = float(raw)
                    except (TypeError, ValueError):
                        v = raw

                cell = ws.cell(row=r, column=2 + i, value=v)
                cell.number_format = fmt
                cell.font      = Font(size=10, color=C_DARK)
                cell.fill      = _fill(bg)
                cell.alignment = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[r].height = 14
            r += 1
        r += 1  # gap between sections

    # Sources-used footer
    all_sources = set()
    for yr_data in reconciled.values():
        all_sources.update(yr_data.get("_sources", {}).values())
    src_labels = {"yahoo": "Yahoo Finance", "edgar": "SEC EDGAR", "macrotrends": "Macrotrends"}
    src_str = "  |  ".join(src_labels.get(s, s) for s in sorted(all_sources))
    r += 1
    fc = ws.cell(row=r, column=1, value=f"Sources used in this model:  {src_str}")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1 + n)
    fc.font      = Font(size=9, italic=True, color="595959")
    fc.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[r].height = 14


# ---------------------------------------------------------------------------
# Tab N — Data Audit  (anomaly flags · FMP cross-check · Companies House)
# ---------------------------------------------------------------------------

def _build_data_audit_sheet(wb: Workbook, valued: dict, meta: dict):
    """Tab: Anomaly flags, FMP cross-check values, and Companies House details."""
    ws = wb.create_sheet("Data Audit")
    ws.sheet_view.showGridLines = False
    _col_width(ws, 1, 14)   # Rule ID / label
    _col_width(ws, 2, 12)   # Severity / value
    _col_width(ws, 3, 22)   # Field
    _col_width(ws, 4, 8)    # Year
    _col_width(ws, 5, 55)   # Issue / detail
    _col_width(ws, 6, 22)   # Observed value / FMP value
    _col_width(ws, 7, 48)   # Recommended action

    SEV_BG    = {"critical": "FCE4D6", "warning": "FFF2CC", "info": "E2EFDA"}
    SEV_FG    = {"critical": "C00000", "warning": "7F6000", "info": "375623"}

    def _fv(v):
        """Format a raw numeric value as B/M/number string."""
        if v is None:
            return "—"
        try:
            v = float(v)
            if abs(v) >= 1e9:
                return f"{v/1e9:.2f}B"
            if abs(v) >= 1e6:
                return f"{v/1e6:.0f}M"
            return f"{v:,.2f}"
        except (TypeError, ValueError):
            return str(v)

    r = 1
    _merge_write(ws, r, 1, 7,
                 f"Data Quality Audit  —  {meta['company']} ({meta['ticker']})",
                 font=Font(bold=True, size=12, color=C_WHITE),
                 fill=_fill(C_NAVY), height=22,
                 align=Alignment(horizontal="left", vertical="center", indent=2))
    r += 2

    # ── Data Quality Score summary ────────────────────────────────────────────
    dq      = valued.get("data_quality", {})
    summary = dq.get("summary", {})
    score   = dq.get("quality_score")

    _section_header(ws, r, 1, 7, "Data Quality Score", color=C_NAVY)
    r += 1

    score_bg = C_GREEN if (score or 0) >= 75 else C_AMBER if (score or 0) >= 50 else C_RED
    score_items = [
        ("Quality Score",          f"{score if score is not None else '—'} / 100",   score_bg),
        ("Status",                 (dq.get("status") or "—").upper(),                 None),
        ("Critical Issues",        summary.get("critical_count", 0),                  None),
        ("Warnings",               summary.get("warning_count",  0),                  None),
        ("Info Notices",           summary.get("info_count",     0),                  None),
        ("Fields needing recheck", ", ".join(summary.get("must_recheck_fields", [])) or "None", None),
    ]
    for label, val, bg in score_items:
        _label_cell(ws, r, 1, label, bold=(label == "Quality Score"))
        vc = ws.cell(row=r, column=2, value=str(val) if val is not None else "—")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        vc.font      = Font(bold=(label == "Quality Score"), size=10, color=C_DARK)
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1,
                                 wrap_text=True)
        if bg:
            vc.fill = _fill(bg)
        r += 1

    # ── Data QA Corrections ───────────────────────────────────────────────────
    qa_log = valued.get("data_qa_log", [])
    r += 1
    _section_header(ws, r, 1, 7,
                    f"Data QA Corrections  ({len(qa_log)} applied pre-pipeline)",
                    color=C_NAVY)
    r += 1

    if not qa_log:
        nc = ws.cell(row=r, column=1,
                     value="No corrections required — all sources agreed within 5% tolerance.")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        nc.font      = Font(size=9, italic=True, color="375623")
        nc.alignment = Alignment(indent=1)
        ws.row_dimensions[r].height = 15
        r += 1
    else:
        for col, txt in [
            (1, "Field"), (2, "Year"), (3, "Original Value"),
            (4, "Original Source"), (5, "Corrected Value"),
            (6, "Corrected Source"), (7, "Reason"),
        ]:
            hc = ws.cell(row=r, column=col, value=txt)
            hc.font      = Font(bold=True, size=9, color=C_WHITE)
            hc.fill      = _fill(C_BLUE)
            hc.alignment = Alignment(
                horizontal="center" if col > 1 else "left",
                vertical="center", indent=1 if col == 1 else 0,
            )
        ws.row_dimensions[r].height = 14
        r += 1

        for corr in qa_log:
            row_bg = C_AMBER   # all corrections are noteworthy
            for col, val in [
                (1, corr.get("field", "—")),
                (2, str(corr.get("year", "—"))),
                (3, _fv(corr.get("original_value"))),
                (4, (corr.get("original_source") or "—").title()),
                (5, _fv(corr.get("corrected_value"))),
                (6, (corr.get("corrected_source") or "—").title()),
                (7, corr.get("reason", "—")),
            ]:
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = Font(
                    bold=(col == 6),
                    size=9,
                    color="375623" if col == 6 else C_DARK,
                )
                cell.fill = _fill(row_bg if col in (5, 6) else
                                  (C_LGRAY if r % 2 == 0 else C_WHITE))
                cell.alignment = Alignment(
                    horizontal="center" if col in (2, 4, 6) else "left",
                    vertical="top",
                    indent=1 if col in (1, 7) else 0,
                    wrap_text=(col == 7),
                )
            reason_len = len(corr.get("reason", ""))
            ws.row_dimensions[r].height = max(14, min(50, 9 * (reason_len // 80 + 1) + 4))
            r += 1

    # ── Anomaly Detector Issues ───────────────────────────────────────────────
    issues = dq.get("issues", [])
    r += 1
    _section_header(ws, r, 1, 7,
                    f"Anomaly Detector Issues  ({len(issues)} total)",
                    color=C_NAVY)
    r += 1

    if not issues:
        nc = ws.cell(row=r, column=1,
                     value="No anomalies detected — data quality checks all passed.")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        nc.font      = Font(size=9, italic=True, color="375623")
        nc.alignment = Alignment(indent=1)
        ws.row_dimensions[r].height = 15
        r += 1
    else:
        # Column headers
        for col, txt in [
            (1, "Rule ID"), (2, "Severity"), (3, "Field"),
            (4, "Year"),    (5, "Issue"),    (6, "Observed Value"),
            (7, "Recommended Action"),
        ]:
            hc = ws.cell(row=r, column=col, value=txt)
            hc.font      = Font(bold=True, size=9, color=C_WHITE)
            hc.fill      = _fill(C_BLUE)
            hc.alignment = Alignment(
                horizontal="center" if col > 1 else "left",
                vertical="center", indent=1 if col == 1 else 0
            )
        ws.row_dimensions[r].height = 14
        r += 1

        ordered = (
            [i for i in issues if i.get("severity") == "critical"] +
            [i for i in issues if i.get("severity") == "warning"]  +
            [i for i in issues if i.get("severity") == "info"]
        )
        for iss in ordered:
            sev = iss.get("severity", "info")
            bg  = SEV_BG.get(sev, C_WHITE)
            fg  = SEV_FG.get(sev, C_DARK)
            yr  = iss.get("year")
            ov  = iss.get("observed_value")

            row_data = [
                (1, iss.get("rule_id",           "—"),         False, False),
                (2, sev.upper(),                               True,  False),
                (3, iss.get("field",             "—"),         False, False),
                (4, str(yr) if yr is not None else "all",      False, False),
                (5, iss.get("message",           "—"),         False, True),
                (6, str(ov) if ov is not None else "—",        False, False),
                (7, iss.get("recommended_action","—"),         False, True),
            ]
            for col, val, bold_col, wrap in row_data:
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = Font(
                    bold=bold_col and sev == "critical",
                    size=9,
                    color=fg if col == 2 else C_DARK,
                )
                cell.fill      = _fill(bg)
                cell.alignment = Alignment(
                    wrap_text=wrap,
                    horizontal="center" if col in (2, 4) else "left",
                    vertical="top",
                    indent=1 if col in (1, 3, 5, 7) else 0,
                )
            msg_len = len(iss.get("message", "")) + len(iss.get("recommended_action", ""))
            ws.row_dimensions[r].height = max(15, min(72, 14 * (msg_len // 100 + 1)))
            r += 1

    # ── FMP Cross-Check ───────────────────────────────────────────────────────
    fmp = valued.get("fmp_crosscheck", {})
    r += 1
    _section_header(ws, r, 1, 7, "FMP Cross-Check Results", color=C_NAVY)
    r += 1

    if not fmp or not fmp.get("available"):
        reason = (fmp or {}).get(
            "reason", "FMP cross-check not run — no flagged fields or FMP_API_KEY not set."
        )
        nc = ws.cell(row=r, column=1, value=reason)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        nc.font      = Font(size=9, italic=True, color="595959")
        nc.alignment = Alignment(indent=1)
        ws.row_dimensions[r].height = 15
        r += 1
    else:
        # Summary meta
        for label, val in [
            ("Fields cross-checked", ", ".join(fmp.get("fields_covered", []))),
            ("API calls used",       fmp.get("calls_made", "—")),
        ]:
            _label_cell(ws, r, 1, label)
            mc = ws.cell(row=r, column=2, value=str(val))
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            mc.font      = Font(size=9, color=C_DARK)
            mc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            r += 1
        r += 1

        fmp_years  = sorted(fmp.get("by_year", {}).keys(), reverse=True)
        fmp_fields = fmp.get("fields_covered", [])

        if fmp_years and fmp_fields:
            # Column headers
            for col, txt in [
                (1, "Field"), (2, "Year"), (3, "FMP Value"),
                (4, "Pipeline Value"), (5, "Difference"),
                (6, "% Diff"), (7, "Status"),
            ]:
                hc = ws.cell(row=r, column=col, value=txt)
                hc.font      = Font(bold=True, size=9, color=C_WHITE)
                hc.fill      = _fill(C_BLUE)
                hc.alignment = Alignment(
                    horizontal="center" if col > 1 else "left",
                    vertical="center",
                )
            ws.row_dimensions[r].height = 14
            r += 1

            reconciled = valued.get("reconciled", {})

            for yr in fmp_years:
                yr_fmp = fmp["by_year"].get(yr, {})
                yr_rec = reconciled.get(yr, {})
                for field in fmp_fields:
                    fv  = yr_fmp.get(field)
                    pv  = yr_rec.get(field)
                    if fv is None and pv is None:
                        continue
                    diff = (fv - pv) if (fv is not None and pv is not None) else None
                    pct  = (abs(diff) / max(abs(pv), 1) * 100) if diff is not None and pv else None
                    ok   = pct is not None and pct <= 5
                    flag = ("MISMATCH" if pct is not None and pct > 5
                            else "OK"  if pct is not None
                            else "—")
                    row_bg = "FCE4D6" if flag == "MISMATCH" else "E2EFDA" if flag == "OK" else C_WHITE

                    for col, val in [
                        (1, field),
                        (2, str(yr)),
                        (3, _fv(fv)),
                        (4, _fv(pv)),
                        (5, _fv(diff)),
                        (6, f"{pct:.1f}%" if pct is not None else "—"),
                        (7, flag),
                    ]:
                        cell = ws.cell(row=r, column=col, value=val)
                        cell.font = Font(
                            bold=(col == 7 and flag == "MISMATCH"),
                            size=9,
                            color=(SEV_FG["critical"] if flag == "MISMATCH" and col == 7
                                   else SEV_FG["info"]    if flag == "OK"      and col == 7
                                   else C_DARK),
                        )
                        cell.fill      = _fill(row_bg if col == 7 else
                                               (C_LGRAY if r % 2 == 0 else C_WHITE))
                        cell.alignment = Alignment(
                            horizontal="center" if col > 1 else "left",
                            vertical="center",
                            indent=1 if col == 1 else 0,
                        )
                    ws.row_dimensions[r].height = 14
                    r += 1

    # ── Companies House ───────────────────────────────────────────────────────
    ch = valued.get("raw", {}).get("companies_house", {})
    r += 1
    _section_header(ws, r, 1, 7,
                    "Companies House  (UK Statutory Registry)", color=C_NAVY)
    r += 1

    if not ch or not ch.get("available"):
        reason = (ch or {}).get("reason", "Companies House data not available.")
        nc = ws.cell(row=r, column=1, value=reason)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        nc.font      = Font(size=9, italic=True, color="595959")
        nc.alignment = Alignment(indent=1)
        ws.row_dimensions[r].height = 15
        r += 1
    else:
        ch_rows = [
            ("Company Number",      ch.get("company_number",     "—")),
            ("Registered Name",     ch.get("registered_name",    "—")),
            ("Status",              ch.get("status",             "—")),
            ("Incorporation Date",  ch.get("incorporation_date", "—")),
            ("SIC Codes",           ", ".join(ch.get("sic_codes", []))),
            ("SIC Descriptions",    ", ".join(ch.get("sic_descriptions", []))),
            ("Last Accounts Date",  ch.get("last_accounts_date", "—")),
            ("Accounts Type",       ch.get("accounts_type",      "—")),
            ("Last Filing Date",    ch.get("last_filing_date",   "—")),
            ("Registered Address",  ch.get("registered_address", "—")),
        ]
        for label, val in ch_rows:
            _label_cell(ws, r, 1, label)
            vc = ws.cell(row=r, column=2, value=str(val) if val else "—")
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            vc.font      = Font(size=9, color=C_DARK)
            vc.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True, indent=1)
            ws.row_dimensions[r].height = 14
            r += 1

        # Notes
        notes = ch.get("notes", [])
        if notes:
            r += 1
            _section_header(ws, r, 1, 7, "Companies House Notes", color=C_BLUE)
            r += 1
            for note in notes:
                is_warn = "WARNING" in note.upper()
                nc = ws.cell(row=r, column=1, value=note)
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
                nc.font = Font(
                    size=9,
                    color="C00000" if is_warn else C_DARK,
                    italic=not is_warn,
                )
                nc.fill      = _fill("FCE4D6" if is_warn else C_LGRAY)
                nc.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
                ws.row_dimensions[r].height = max(
                    15, min(50, 9 * (len(note) // 90 + 1) + 4)
                )
                r += 1

    # ── Source Comparison for Flagged Fields ──────────────────────────────────
    # For every critical / warning flag, show what each source actually returned.
    # This lets the analyst see where sources disagree and which value was used.
    raw_data    = valued.get("raw", {})
    reconciled2 = valued.get("reconciled", {})
    fmp2        = valued.get("fmp_crosscheck", {})

    yahoo_by_yr = (raw_data.get("yahoo") or {}).get("financials_by_year", {})
    edgar_info  = raw_data.get("edgar") or {}
    edgar_fin   = edgar_info.get("financials", {}) if edgar_info.get("available") else {}
    edgar_yr    = str(edgar_info.get("latest_year", ""))
    mt_info     = raw_data.get("macrotrends") or {}
    mt_by_yr    = mt_info.get("financials_by_year", {}) if mt_info.get("available") else {}
    fmp_by_yr2  = fmp2.get("by_year", {}) if (fmp2 or {}).get("available") else {}

    # Collect unique (severity, field, year) from critical + warning issues
    seen_pairs: set = set()
    flagged_rows: list = []
    for iss in sorted(
        [i for i in dq.get("issues", []) if i.get("severity") in ("critical", "warning")],
        key=lambda x: (0 if x.get("severity") == "critical" else 1,
                       x.get("field", ""), str(x.get("year", ""))),
    ):
        f   = iss.get("field")
        raw_yr = iss.get("year")
        if not f:
            continue
        if raw_yr is None:
            # Issue applies to all years — show most recent 5 where the field exists
            expand = sorted(
                [y for y in reconciled2 if reconciled2[y].get(f) is not None],
                reverse=True,
            )[:5]
        else:
            expand = [str(raw_yr)]
        for yr in expand:
            key = (f, str(yr))
            if key not in seen_pairs:
                seen_pairs.add(key)
                flagged_rows.append((iss.get("severity", "info"), f, str(yr)))

    r += 2
    _col_width(ws, 8, 14)
    n_cols_sc = 8
    _section_header(ws, r, 1, n_cols_sc,
                    f"Source Comparison — Flagged Fields  ({len(flagged_rows)} rows)",
                    color=C_NAVY)
    r += 1

    sub = ws.cell(row=r, column=1,
                  value="Shows what each data source returned for every flagged field/year. "
                        "'Reconciled' column shows the value used in the model and its source.")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols_sc)
    sub.font      = Font(size=9, italic=True, color="595959")
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[r].height = 14
    r += 1

    if not flagged_rows:
        nc = ws.cell(row=r, column=1, value="No critical or warning flags to compare.")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols_sc)
        nc.font      = Font(size=9, italic=True, color="595959")
        nc.alignment = Alignment(indent=1)
        ws.row_dimensions[r].height = 15
    else:
        for col, txt in [
            (1, "Field"), (2, "Year"), (3, "Yahoo"),
            (4, "EDGAR"), (5, "Macrotrends"), (6, "FMP"),
            (7, "Reconciled (used)"), (8, "Severity"),
        ]:
            hc = ws.cell(row=r, column=col, value=txt)
            hc.font      = Font(bold=True, size=9, color=C_WHITE)
            hc.fill      = _fill(C_BLUE)
            hc.alignment = Alignment(
                horizontal="center" if col > 1 else "left",
                vertical="center", indent=1 if col == 1 else 0,
            )
        ws.row_dimensions[r].height = 14
        r += 1

        for sev, field, yr in flagged_rows:
            yf_val  = _fv(yahoo_by_yr.get(yr, {}).get(field))
            ed_val  = _fv(edgar_fin.get(field)) if yr == edgar_yr else "—"
            mt_val  = _fv(mt_by_yr.get(yr, {}).get(field))
            fmp_val = _fv(fmp_by_yr2.get(yr, {}).get(field)) if fmp_by_yr2 else "—"
            rec_raw = reconciled2.get(yr, {}).get(field)
            src_used = (reconciled2.get(yr, {}).get("_sources", {}).get(field) or "").title()
            rec_str  = f"{_fv(rec_raw)}  [{src_used}]" if src_used else _fv(rec_raw)

            bg = SEV_BG.get(sev, C_WHITE)
            fg = SEV_FG.get(sev, C_DARK)

            for col, val in [
                (1, field), (2, yr),     (3, yf_val),
                (4, ed_val),(5, mt_val), (6, fmp_val),
                (7, rec_str), (8, sev.upper()),
            ]:
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = Font(
                    bold=(col == 8 and sev == "critical"),
                    size=9,
                    color=fg if col == 8 else C_DARK,
                )
                cell.fill = _fill(
                    bg if col in (7, 8) else
                    (C_LGRAY if r % 2 == 0 else C_WHITE)
                )
                cell.alignment = Alignment(
                    horizontal=(
                        "center" if col in (2, 8) else
                        "right"  if col in (3, 4, 5, 6, 7) else
                        "left"
                    ),
                    vertical="center",
                    indent=1 if col == 1 else 0,
                )
            ws.row_dimensions[r].height = 14
            r += 1


# ---------------------------------------------------------------------------
# Tab N+1 — Pipeline & Flags
# ---------------------------------------------------------------------------

def _build_pipeline_sheet(wb: Workbook, valued: dict, meta: dict):
    ws = wb.create_sheet("Pipeline & Flags")
    ws.sheet_view.showGridLines = False
    _col_width(ws, 1, 28)
    _col_width(ws, 2, 16)
    _col_width(ws, 3, 14)
    _col_width(ws, 4, 72)

    r = 1
    _merge_write(ws, r, 1, 3,
                 f"Pipeline Status  —  {meta['company']} ({meta['ticker']})",
                 font=Font(bold=True, size=12, color=C_WHITE),
                 fill=_fill(C_NAVY), height=22,
                 align=Alignment(horizontal="left", vertical="center", indent=2))

    r += 2
    _section_header(ws, r, 1, 3, "Stage Status Summary")
    r += 1

    stages = [
        ("01 Scrape",            "scrape",             "pass"),
        ("02 Standardise",       "standardise",        "pass"),
        ("03 Validate",          "validate",           valued.get("quality_score", 0)),
        ("04 Normalise",         "normalise",          valued.get("normalised", {}).get("ebit_margin")),
        ("05 Classify",          "classify",           valued.get("classification", "")),
        ("06 Assumption Engine", "assumption_engine",  len(valued.get("assumptions", {}))),
        ("07 Forecast",          "forecast",           len(valued.get("forecast", {}))),
        ("08 Valuation Engine",  "valuation_engine",   valued.get("valuation", {}).get("base", {}).get("value_per_share")),
    ]

    for label, _, indicator in stages:
        ok = indicator not in (None, 0, "", {}, [])
        _label_cell(ws, r, 1, label, bold=True)
        status_cell = ws.cell(row=r, column=2, value="PASS" if ok else "INCOMPLETE")
        status_cell.font = Font(bold=True, size=10,
                                color="375623" if ok else "C00000")
        status_cell.alignment = Alignment(horizontal="center")
        r += 1

    r += 1
    _section_header(ws, r, 1, 3, "Key Model Metadata")
    r += 1

    meta_rows = [
        ("Company",          meta['company']),
        ("Ticker",           meta['ticker']),
        ("Classification",   valued.get("classification", "—")),
        ("Template",         valued.get("template", "—")),
        ("Base Year",        valued.get("base_year", "—")),
        ("Data Quality Score", f"{valued.get('quality_score', '—')}  / 100"),
        ("Run Date",         meta['date']),
        ("Currency",         valued.get("stats", {}).get("currency", "—")),
        ("Country",          valued.get("stats", {}).get("country", "—")),
        ("Sector",           valued.get("stats", {}).get("sector", "—")),
        ("Industry",         valued.get("stats", {}).get("industry", "—")),
    ]
    for label, val in meta_rows:
        _label_cell(ws, r, 1, label)
        cell = ws.cell(row=r, column=2, value=str(val) if val else "—")
        cell.font = Font(size=10, color=C_DARK)
        r += 1

    r += 1
    _section_header(ws, r, 1, 3, "Warnings & Flags from Pipeline")
    r += 1

    all_warnings = valued.get("warnings", [])
    if not all_warnings:
        _label_cell(ws, r, 1, "No warnings recorded.")
        r += 1
    for w in all_warnings:
        cell = ws.cell(row=r, column=1, value=w)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        cell.font = Font(size=9, color="595959", italic=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 28
        r += 1

    # -----------------------------------------------------------------------
    # Coherence Check Results
    # -----------------------------------------------------------------------
    coherence = valued.get("coherence", {})
    checks    = coherence.get("checks", [])

    if checks:
        r += 1
        coh_status = coherence.get("status", "")
        coh_label  = {
            "review_required": "REVIEW REQUIRED",
            "caution":         "CAUTION",
            "pass":            "ALL CHECKS PASSED",
        }.get(coh_status, coh_status.upper())
        coh_color  = {
            "review_required": "C00000",
            "caution":         C_WARN,
            "pass":            "375623",
        }.get(coh_status, C_DARK)

        n_flags = len(coherence.get("flags", []))
        n_warns = len(coherence.get("warns", []))
        n_pass  = len(coherence.get("passes", []))
        header_text = (
            f"Coherence Checks  —  {coh_label}  "
            f"({n_flags} flag{'s' if n_flags!=1 else ''}  ·  "
            f"{n_warns} warning{'s' if n_warns!=1 else ''}  ·  "
            f"{n_pass} passed)"
        )
        _section_header(ws, r, 1, 4, header_text, color=C_NAVY)
        _col_width(ws, 4, 70)   # detail column — wide
        ws.row_dimensions[r].height = 17
        r += 1

        # Column header row
        for col, txt in [(1, "Check"), (2, "Status"), (3, "Value"), (4, "Detail")]:
            cell = ws.cell(row=r, column=col, value=txt)
            cell.font  = Font(bold=True, size=9, color=C_WHITE)
            cell.fill  = _fill(C_BLUE)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 14
        r += 1

        STATUS_COLOR = {"flag": "C00000", "warn": "ED7D31", "pass": "375623"}
        STATUS_BG    = {"flag": "FCE4D6", "warn": "FFF2CC", "pass": "E2EFDA"}
        STATUS_LABEL = {"flag": "FLAG", "warn": "WARN", "pass": "PASS"}

        # Sort: flags first, then warns, then passes
        ordered = (
            [c for c in checks if c["status"] == "flag"] +
            [c for c in checks if c["status"] == "warn"] +
            [c for c in checks if c["status"] == "pass"]
        )

        for chk in ordered:
            st   = chk["status"]
            bg   = STATUS_BG.get(st, C_WHITE)
            fc   = STATUS_COLOR.get(st, C_DARK)
            lbl  = STATUS_LABEL.get(st, st.upper())

            # Title
            tc = ws.cell(row=r, column=1, value=chk["title"])
            tc.font      = Font(bold=True, size=9, color=C_DARK)
            tc.fill      = _fill(bg)
            tc.alignment = Alignment(vertical="top", indent=1)

            # Status badge
            sc = ws.cell(row=r, column=2, value=lbl)
            sc.font      = Font(bold=True, size=9, color=fc)
            sc.fill      = _fill(bg)
            sc.alignment = Alignment(horizontal="center", vertical="top")

            # Value
            vc = ws.cell(row=r, column=3, value=chk["value"])
            vc.font      = Font(size=9, color=C_DARK)
            vc.fill      = _fill(bg)
            vc.alignment = Alignment(horizontal="center", vertical="top")

            # Detail (merged across col 4 only — wide column handles wrapping)
            dc = ws.cell(row=r, column=4, value=chk["detail"])
            dc.font      = Font(size=9, color=C_DARK,
                                bold=(st == "flag"), italic=(st == "pass"))
            dc.fill      = _fill(bg)
            dc.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
            ws.row_dimensions[r].height = max(
                28, min(80, 14 * (len(chk["detail"]) // 80 + 1))
            )
            r += 1

        # Assumption confidence score summary row
        conf = coherence.get("assumption_confidence_score")
        if conf is not None:
            r += 1
            conf_bg = C_GREEN if conf >= 75 else C_AMBER if conf >= 50 else C_RED
            cell = ws.cell(row=r, column=1,
                           value=f"Assumption Confidence Score:  {conf} / 100")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            cell.font      = Font(bold=True, size=9, color=C_DARK)
            cell.fill      = _fill(conf_bg)
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            ws.row_dimensions[r].height = 15
            r += 1

    # -----------------------------------------------------------------------
    # Review Agenda — full detail of every flag and warning
    # -----------------------------------------------------------------------
    expl        = valued.get("explanation", {})
    agenda      = expl.get("review_agenda", [])
    pipeline_ws = valued.get("warnings", [])
    norm_warns  = valued.get("normalised", {}).get("warnings", []) if valued.get("normalised") else []

    r += 1
    _section_header(ws, r, 1, 4, "REVIEW AGENDA  —  All items requiring analyst attention",
                    color=C_NAVY)
    ws.row_dimensions[r].height = 17
    r += 1

    if not agenda or (len(agenda) == 1 and agenda[0].get("status") == "PASS"):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        cell = ws.cell(row=r, column=1, value="No items flagged — all coherence checks passed.")
        cell.font      = Font(italic=True, size=9, color="375623")
        cell.alignment = Alignment(indent=1)
        ws.row_dimensions[r].height = 15
        r += 1
    else:
        # Column headers
        for col, txt in [(1, "Item"), (2, "Status"), (3, "Value"), (4, "Action Required")]:
            cell = ws.cell(row=r, column=col, value=txt)
            cell.font      = Font(bold=True, size=9, color=C_WHITE)
            cell.fill      = _fill(C_BLUE)
            cell.alignment = Alignment(horizontal="center" if col > 1 else "left",
                                       vertical="center", indent=1 if col == 1 else 0)
        ws.row_dimensions[r].height = 14
        r += 1

        ST_BG  = {"FLAG": "FCE4D6", "WARN": "FFF2CC", "PASS": "E2EFDA"}
        ST_FG  = {"FLAG": "C00000", "WARN": "7F6000", "PASS": "375623"}

        for item in agenda:
            st  = item.get("status", "PASS")
            if st == "PASS":
                continue
            bg  = ST_BG.get(st, C_WHITE)
            fg  = ST_FG.get(st, C_DARK)

            # Priority + item name
            nc = ws.cell(row=r, column=1,
                         value=f"{item['priority']}.  {item['item']}")
            nc.font      = Font(bold=(st == "FLAG"), size=9, color=fg)
            nc.fill      = _fill(bg)
            nc.alignment = Alignment(vertical="top", indent=1)

            # Status badge
            sc = ws.cell(row=r, column=2, value=st)
            sc.font      = Font(bold=True, size=9, color=fg)
            sc.fill      = _fill(bg)
            sc.alignment = Alignment(horizontal="center", vertical="top")

            # Observed value
            vc = ws.cell(row=r, column=3, value=item.get("value", "—"))
            vc.font      = Font(size=9, color=C_DARK)
            vc.fill      = _fill(bg)
            vc.alignment = Alignment(horizontal="center", vertical="top")

            # Action text — full detail
            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=4)
            ac = ws.cell(row=r, column=4, value=item.get("action", ""))
            ac.font      = Font(bold=(st == "FLAG"), size=9, color=C_DARK)
            ac.fill      = _fill(bg)
            ac.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
            action_len   = len(item.get("action", ""))
            ws.row_dimensions[r].height = max(28, min(90, 9 * (action_len // 72 + 1) + 6))
            r += 1

    # -----------------------------------------------------------------------
    # All pipeline warnings (from normaliser, validator, etc.)
    # -----------------------------------------------------------------------
    all_pipe_warns = valued.get("warnings", [])
    if all_pipe_warns:
        r += 1
        _section_header(ws, r, 1, 4,
                        "PIPELINE WARNINGS  —  From data processing stages",
                        color=C_BLUE)
        ws.row_dimensions[r].height = 17
        r += 1
        for i, w in enumerate(all_pipe_warns):
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            cell = ws.cell(row=r, column=1, value=w)
            cell.font      = Font(size=9, color="595959", italic=True)
            cell.fill      = _fill(C_LGRAY if i % 2 == 0 else C_WHITE)
            cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
            ws.row_dimensions[r].height = max(15, min(60, 9 * (len(w) // 90 + 1) + 4))
            r += 1


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def _extract_meta(valued: dict) -> dict:
    company = valued.get("company_name", "Unknown")
    ticker  = valued.get("ticker", "")
    safe    = "".join(c if c.isalnum() or c in "._- " else "_" for c in company)[:40]
    return {
        "company":   company,
        "ticker":    ticker,
        "safe_name": safe.replace(" ", "_"),
        "date":      datetime.utcnow().strftime("%Y-%m-%d"),
    }


def _m(val) -> float | None:
    """Convert to millions for display."""
    if val is None:
        return None
    try:
        return float(val) / _DIV
    except (TypeError, ValueError):
        return None


def _smart_m(val):
    """Scale value adaptively: returns (display_value, format_string).
    Uses B suffix when |raw value| >= 1 billion, otherwise M."""
    if val is None:
        return None, FMT_M
    try:
        v = float(val)
    except (TypeError, ValueError):
        return None, FMT_M
    if abs(v) >= 1_000_000_000:
        return v / 1_000_000_000, FMT_B
    return v / 1_000_000, FMT_M


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _top_border() -> Border:
    s = Side(style="thin", color="1F3864")
    return Border(top=s)


def _col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


def _merge_write(ws, r, c1, c2, text, font, fill, height, align):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c1, value=text)
    cell.font   = font
    cell.fill   = fill
    cell.alignment = align
    ws.row_dimensions[r].height = height


def _section_header(ws, r, c1, c2, text, color=C_NAVY):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c1, value=text)
    cell.font  = Font(bold=True, size=10, color=C_WHITE)
    cell.fill  = _fill(color)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 17


def _header_cell(ws, r, c, text, bg=C_BLUE):
    cell = ws.cell(row=r, column=c, value=text)
    cell.font  = Font(bold=True, size=10, color=C_WHITE if bg not in (C_HIST, C_FORE, C_GREEN, C_RED, C_LGRAY) else C_DARK)
    cell.fill  = _fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 16


def _label_cell(ws, r, c, text, bold=False, italic=False, indent=1):
    cell = ws.cell(row=r, column=c, value=text)
    cell.font  = Font(bold=bold, italic=italic, size=10, color=C_DARK)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
    ws.row_dimensions[r].height = 15


def _input_cell(ws, r, c, value, fmt="0.0%", bold=False, bg=None):
    cell = ws.cell(row=r, column=c, value=value)
    cell.font  = Font(bold=bold, size=10, color=C_INPUT)
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if bg:
        cell.fill = _fill(bg)
    ws.row_dimensions[r].height = 15


def _calc_cell(ws, r, c, value, fmt=FMT_M, bold=False,
               bg=None, color=None, italic=False, border=None):
    cell = ws.cell(row=r, column=c, value=value)
    cell.font  = Font(bold=bold, italic=italic, size=10,
                      color=color or C_DARK)
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if bg:
        cell.fill = _fill(bg)
    if border:
        cell.border = border
    ws.row_dimensions[r].height = 15
